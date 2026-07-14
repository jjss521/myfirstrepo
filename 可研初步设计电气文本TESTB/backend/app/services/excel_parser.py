"""
Excel负荷计算解析器

解析给水/排水/环卫工程常用的"需要系数法"负荷计算Excel文件。
支持 .xlsx 和 .xls 两种格式。

Excel文件结构（已从实际文件分析确认）：
- 列结构: 设备组名称 | 额定功率(kW) | 安装台数 | 工作台数 | 设备功率(kW) | 
          需要系数(KX) | cosφ | tanφ | 无功补偿率qc | PC(kW) | QC(kvar) | SC(kVA) | 备注
- 按区域分组（如取水泵房、送水泵房等）
- 有"用电设备组计算负荷"和"总计算负荷"汇总行
- 包含补偿前后功率因数计算
- 可选"需要系数Kx+cosφ"参考表sheet
"""
import re
import math
from typing import List, Dict, Any, Optional, Tuple


class ExcelLoadParser:
    """负荷计算Excel解析器"""

    # 需要系数参考表（来自实际文件 + 电气设计手册）
    KX_COS_REFERENCE = {
        '取水泵': (0.9, 0.85),
        '加压泵': (0.9, 0.85),
        '污水提升泵': (0.9, 0.85),
        '送水泵': (0.9, 0.85),
        '进水泵': (0.9, 0.8),
        '冲洗泵': (0.7, 0.85),
        '真空泵': (0.5, 0.8),
        '排水泵': (0.3, 0.8),
        '鼓风机': (0.7, 0.85),
        '通风机': (0.7, 0.85),
        '轴流风机': (0.7, 0.85),
        '搅拌机': (0.8, 0.8),
        '搅拌器': (0.8, 0.8),
        '刮泥机': (0.8, 0.8),
        '投药': (0.7, 0.8),
        '加药': (0.8, 0.8),
        '消毒': (0.9, 0.5),
        '电动阀门': (0.2, 0.8),
        '电动闸阀': (0.2, 0.8),
        '电动闸门': (0.2, 0.8),
        '起重机': (0.2, 0.5),
        '电动葫芦': (0.2, 0.8),
        '污泥脱水': (0.7, 0.8),
        '格栅': (0.6, 0.75),
        '除臭': (0.7, 0.8),
        '照明': (0.8, 0.9),
        '厂区照明': (0.9, 0.9),
        '自控仪表': (0.2, 0.7),
        '仪表': (0.2, 0.7),
        '化验室': (0.5, 0.9),
        '机修': (0.4, 0.8),
        '仓库': (0.4, 0.8),
        'PLC': (0.8, 0.85),
        '计算机': (0.5, 0.5),
    }

    # 汇总行识别关键词
    SUMMARY_KEYWORDS = ['用电设备组计算负荷', '总计算负荷', '计算负荷', '合计']
    AREA_KEYWORDS = ['泵房', '车间', '间', '池', '站', '系统', '建筑', '房']
    SKIP_KEYWORDS = ['同时系数', '补偿前', '补偿后', '要求补偿', '电力电容器']

    def __init__(self):
        self.raw_rows = []
        self.areas = {}  # {area_name: [devices]}
        self.summary = {}  # 全厂汇总

    def parse(self, file_path: str) -> Dict[str, Any]:
        """
        主解析入口
        返回结构化解析结果
        """
        ext = file_path.lower()
        if ext.endswith('.xlsx'):
            self._parse_xlsx(file_path)
        elif ext.endswith('.xls'):
            self._parse_xls(file_path)
        else:
            raise ValueError(f'不支持的文件格式: {ext}')

        return self._build_result()

    def _parse_xlsx(self, file_path: str):
        """解析 .xlsx 文件"""
        try:
            import openpyxl
        except ImportError:
            raise ImportError('需要安装 openpyxl: pip install openpyxl')

        wb = openpyxl.load_workbook(file_path, data_only=True)
        # 优先使用第一个非空sheet（跳过Kx参考表等）
        main_sheets = [s for s in wb.sheetnames
                       if '需要系数' not in s and '变压器' not in s]
        if not main_sheets:
            main_sheets = wb.sheetnames

        for sheet_name in main_sheets:
            ws = wb[sheet_name]
            self._parse_worksheet(ws)

    def _parse_xls(self, file_path: str):
        """解析 .xls 文件"""
        try:
            import xlrd
        except ImportError:
            raise ImportError('需要安装 xlrd: pip install xlrd')

        wb = xlrd.open_workbook(file_path)
        main_sheets = [s for s in wb.sheet_names()
                       if '需要系数' not in s and '变压器' not in s]
        if not main_sheets:
            main_sheets = wb.sheet_names

        for sheet_name in main_sheets:
            ws = wb.sheet_by_name(sheet_name)
            self._parse_worksheet(ws, is_xlrd=True)

    def _clean_value(self, val) -> Optional[str]:
        """清理单元格值"""
        if val is None:
            return None
        s = str(val).strip()
        return s if s else None

    def _is_numeric_val(self, val) -> bool:
        """判断是否为可转换为数值的内容"""
        if val is None:
            return False
        if isinstance(val, (int, float)):
            return True
        try:
            float(str(val).strip())
            return True
        except (ValueError, TypeError):
            return False

    def _to_float(self, val) -> Optional[float]:
        """安全转换为浮点数"""
        if val is None:
            return None
        if isinstance(val, (int, float)):
            return float(val)
        try:
            return float(str(val).strip())
        except (ValueError, TypeError):
            return None

    def _parse_worksheet(self, ws, is_xlrd=False):
        """解析单个工作表"""
        # 转换为二维list便于处理
        rows_data = []
        nrows = ws.nrows if is_xlrd else ws.max_row
        ncols = ws.ncols if is_xlrd else ws.max_column

        for r in range(nrows):
            row_vals = []
            for c in range(min(ncols, 20)):
                if is_xlrd:
                    val = ws.cell_value(r, c)
                else:
                    cell = ws.cell(row=r + 1, column=c + 1)
                    val = cell.value
                row_vals.append(val)
            rows_data.append(row_vals)

        # 解析行数据
        current_area = None
        devices_in_area = []

        for row in rows_data:
            col_a = self._clean_value(row[0]) if len(row) > 0 else None
            if not col_a:
                continue

            # 跳过汇总行和标题行
            if any(kw in col_a for kw in self.SKIP_KEYWORDS):
                continue

            # 检查是否为汇总行
            is_summary = any(kw in col_a for kw in self.SUMMARY_KEYWORDS)

            # 检查是否为区域标题行（无功率数据只有名称的行）
            has_power_data = any(
                self._is_numeric_val(row[c])
                for c in [1, 4] if len(row) > c
            ) if len(row) > 5 else False

            # 判断是否为新的区域分组
            if not has_power_data and not is_summary:
                # 可能是区域标题
                if any(kw in col_a for kw in self.AREA_KEYWORDS):
                    # 保存上一个区域
                    if current_area and devices_in_area:
                        self.areas[current_area] = devices_in_area
                    current_area = col_a
                    devices_in_area = []
                    continue

            # 解析设备行
            if has_power_data:
                device = self._parse_device_row(row, col_a)
                if device:
                    if current_area:
                        devices_in_area.append(device)
                    else:
                        # 无区域的设备直接放全局
                        if '_global' not in self.areas:
                            self.areas['_global'] = []
                        self.areas['_global'].append(device)
                    self.raw_rows.append(device)

            # 解析汇总行
            if is_summary:
                summary_data = self._parse_summary_row(row, col_a)
                if summary_data:
                    if current_area and devices_in_area:
                        self.areas[current_area] = devices_in_area
                        if '_summaries' not in self.areas:
                            self.areas['_summaries'] = {}
                        self.areas['_summaries'][current_area] = summary_data

        # 保存最后一个区域
        if current_area and devices_in_area:
            self.areas[current_area] = devices_in_area

    def _parse_device_row(self, row: list, name: str) -> Optional[Dict]:
        """解析单个设备行"""
        # 提取各列数据（位置: 0名称, 1额定功率, 2安装台数, 3工作台数, 4设备功率, 5Kx, 6cosφ, 7tanφ, 8qc, 9PC, 10QC, 11SC, 12备注）

        rated_power = self._to_float(row[1]) if len(row) > 1 else None  # 单台功率
        install_num = self._to_float(row[2]) if len(row) > 2 else None
        work_num = self._to_float(row[3]) if len(row) > 3 else None
        equip_power = self._to_float(row[4]) if len(row) > 4 else None  # 设备总功率
        kx = self._to_float(row[5]) if len(row) > 5 else None
        cos_phi = self._to_float(row[6]) if len(row) > 6 else None
        tan_phi = self._to_float(row[7]) if len(row) > 7 else None
        qc = self._to_float(row[8]) if len(row) > 8 else None
        pc = self._to_float(row[9]) if len(row) > 9 else None  # 计算有功
        qc_val = self._to_float(row[10]) if len(row) > 10 else None  # 计算无功
        sc = self._to_float(row[11]) if len(row) > 11 else None  # 计算视在
        remark = self._clean_value(row[12]) if len(row) > 12 else None

        # 尝试从名称匹配Kx和cosφ
        if kx is None or cos_phi is None:
            matched_kx, matched_cos = self._match_kx_cos(name)
            if kx is None:
                kx = matched_kx
            if cos_phi is None:
                cos_phi = matched_cos

        # 计算缺失值
        if equip_power is None and rated_power is not None:
            if install_num is not None:
                equip_power = rated_power * install_num

        if pc is None and equip_power is not None and kx is not None:
            pc = equip_power * kx

        if tan_phi is None and cos_phi is not None and 0 < cos_phi <= 1.0:
            tan_phi = math.tan(math.acos(min(cos_phi, 1.0)))

        if qc_val is None and pc is not None and tan_phi is not None:
            qc_val = pc * tan_phi

        if sc is None and pc is not None and qc_val is not None:
            sc = math.sqrt(pc ** 2 + qc_val ** 2)

        # 计算电流 (380V)
        current = None
        if sc is not None:
            current = sc / (math.sqrt(3) * 0.38)

        return {
            'name': name,
            'rated_power': rated_power,  # 单台功率kW
            'install_num': install_num,
            'work_num': work_num,
            'equip_power': equip_power,  # 总设备功率kW
            'kx': kx,
            'cos_phi': cos_phi,
            'tan_phi': tan_phi,
            'qc_rate': qc,
            'pc': pc,  # 计算有功kW
            'qc': qc_val,  # 计算无功kvar
            'sc': sc,  # 计算视在kVA
            'current': current,  # 计算电流A
            'remark': remark,
        }

    def _parse_summary_row(self, row: list, name: str) -> Optional[Dict]:
        """解析汇总行"""
        pc = self._to_float(row[9]) if len(row) > 9 else None
        qc = self._to_float(row[10]) if len(row) > 10 else None
        sc = self._to_float(row[11]) if len(row) > 11 else None
        equip_power = self._to_float(row[4]) if len(row) > 4 else None

        return {
            'name': name,
            'equip_power': equip_power,
            'pc': pc,
            'qc': qc,
            'sc': sc,
        }

    def _match_kx_cos(self, name: str) -> Tuple[Optional[float], Optional[float]]:
        """根据设备名称模糊匹配需要系数和功率因数"""
        name_lower = name.strip()
        for key, (kx, cos) in self.KX_COS_REFERENCE.items():
            if key in name_lower:
                return (kx, cos)
        return (None, None)

    def _build_result(self) -> Dict[str, Any]:
        """构建最终解析结果"""
        # 汇总各区域
        all_devices = []
        area_summaries = {}
        total_equip_power = 0
        total_pc = 0
        total_qc = 0

        for area_name, devices in self.areas.items():
            if area_name.startswith('_'):
                continue
            area_equip = sum(d.get('equip_power', 0) or 0 for d in devices)
            area_pc = sum(d.get('pc', 0) or 0 for d in devices)
            area_qc = sum(d.get('qc', 0) or 0 for d in devices)
            area_sc = math.sqrt(area_pc ** 2 + area_qc ** 2) if area_pc and area_qc else 0

            area_summaries[area_name] = {
                'device_count': len(devices),
                'equip_power': round(area_equip, 2),
                'pc': round(area_pc, 2),
                'qc': round(area_qc, 2),
                'sc': round(area_sc, 2),
            }

            total_equip_power += area_equip
            total_pc += area_pc
            total_qc += area_qc
            all_devices.extend(devices)

        # 全厂汇总（考虑同时系数 KΣP=0.9, KΣq=0.95）
        ksp = 0.9
        ksq = 0.95
        total_pc_k = total_pc * ksp
        total_qc_k = total_qc * ksq
        total_sc_k = math.sqrt(total_pc_k ** 2 + total_qc_k ** 2)

        # 补偿前功率因数
        cos_before = total_pc_k / total_sc_k if total_sc_k > 0 else 0

        # 需要补偿到0.95
        cos_target = 0.95
        cos_before_clamped = min(max(cos_before, 0.01), 1.0)
        qc_compensation = total_pc_k * (math.tan(math.acos(cos_before_clamped)) - math.tan(math.acos(cos_target))) if cos_before > 0 else 0

        total_qc_after = max(0, total_qc_k - qc_compensation)
        total_sc_after = math.sqrt(total_pc_k ** 2 + total_qc_after ** 2)

        self.summary = {
            'total_devices': len(all_devices),
            'total_equip_power': round(total_equip_power, 2),
            'total_pc': round(total_pc, 2),
            'total_qc': round(total_qc, 2),
            'total_sc': round(math.sqrt(total_pc ** 2 + total_qc ** 2) if total_pc else 0, 2),
            'total_pc_k': round(total_pc_k, 2),
            'total_qc_k': round(total_qc_k, 2),
            'total_sc_k': round(total_sc_k, 2),
            'cos_before': round(cos_before, 4),
            'cos_target': cos_target,
            'qc_compensation': round(qc_compensation, 2),
            'total_qc_after': round(total_qc_after, 2),
            'total_sc_after': round(total_sc_after, 2),
            'simultaneous_coeff': {'KΣP': ksp, 'KΣq': ksq},
        }

        # 推荐变压器容量
        recommended_tx = math.ceil(total_sc_after / 100) * 100
        # 推荐变压器台数
        if recommended_tx <= 1600:
            tx_config = f"1×{self._std_tx_capacity(recommended_tx)}kVA"
        elif recommended_tx <= 3200:
            half = math.ceil(recommended_tx / 2 / 50) * 50
            tx_config = f"2×{self._std_tx_capacity(half)}kVA"
        else:
            n = min(4, math.ceil(recommended_tx / 2000))
            each = math.ceil(recommended_tx / n / 50) * 50
            tx_config = f"{n}×{self._std_tx_capacity(each)}kVA"

        self.summary['recommended_transformer'] = tx_config

        return {
            'area_summaries': area_summaries,
            'all_devices': all_devices,
            'summary': self.summary,
            'area_count': len([k for k in self.areas if not k.startswith('_')]),
            'total_areas': len(self.areas),
        }

    def _std_tx_capacity(self, capacity: float) -> int:
        """取标准变压器容量"""
        standard = [100, 160, 200, 250, 315, 400, 500, 630, 800, 1000, 1250, 1600, 2000, 2500]
        for s in standard:
            if s >= capacity:
                return s
        return int(math.ceil(capacity / 100) * 100)


# 快速测试入口
if __name__ == '__main__':
    import json
    parser = ExcelLoadParser()
    result = parser.parse(r"D:\00-水厂负荷计算表.xlsx")
    print(json.dumps(result['summary'], ensure_ascii=False, indent=2))
    print('\n=== 区域汇总 ===')
    for area, data in result['area_summaries'].items():
        print(f"{area}: 设备{data['device_count']}台, 设备功率{data['equip_power']}kW, 计算负荷{data['sc']}kVA")
