# -*- coding: utf-8 -*-
"""Excel报表导出模块 - 生成符合模板格式的负荷计算Excel文件"""

from math import sqrt
from typing import List

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from .calc_engine import calc_subsystem_summary, calc_hv_system_summary
from .models import HVSystem

# ── 样式定义 ──
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

HEADER_FONT = Font(name='微软雅黑', size=9, bold=True)
TITLE_FONT = Font(name='微软雅黑', size=12, bold=True)
SECTION_FONT = Font(name='微软雅黑', size=10, bold=True)
NORMAL_FONT = Font(name='微软雅黑', size=9)
BOLD_FONT = Font(name='微软雅黑', size=9, bold=True)
LABEL_FONT = Font(name='微软雅黑', size=9, bold=True)

HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT_WHITE = Font(name='微软雅黑', size=9, bold=True, color='FFFFFF')
TITLE_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
SECTION_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
SUBTOTAL_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center')
RIGHT = Alignment(horizontal='right', vertical='center')

NUM_FMT_2D = '#,##0.00'
NUM_FMT_1D = '#,##0.0'
NUM_FMT_INT = '#,##0'
NUM_FMT_PCT = '0.0%'
NUM_FMT_PF = '0.0000'


def _apply_cell(ws, row, col, value, font=None, fill=None,
                alignment=None, number_format=None, border=None):
    """设置单元格的值和样式"""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if number_format:
        cell.number_format = number_format
    if border:
        cell.border = border
    return cell


def _write_title_row(ws, row, text, max_col):
    """写入合并标题行"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    _apply_cell(ws, row, 1, text,
                font=TITLE_FONT, fill=TITLE_FILL,
                alignment=CENTER)
    ws.row_dimensions[row].height = 32


def _write_section_header(ws, row, text, max_col):
    """写入节标题"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    _apply_cell(ws, row, 1, text,
                font=SECTION_FONT, fill=SECTION_FILL,
                alignment=LEFT)
    ws.row_dimensions[row].height = 24


def _write_table_header(ws, row, headers):
    """写入表格表头行"""
    for col_idx, h in enumerate(headers, 1):
        _apply_cell(ws, row, col_idx, h,
                    font=HEADER_FONT_WHITE, fill=HEADER_FILL,
                    alignment=CENTER, border=THIN_BORDER)
    ws.row_dimensions[row].height = 22


def _write_data_row(ws, row, values, formats=None, fill=None):
    """写入数据行"""
    for col_idx, v in enumerate(values, 1):
        fmt = formats[col_idx - 1] if formats else None
        _apply_cell(ws, row, col_idx, v,
                    font=NORMAL_FONT, alignment=CENTER,
                    number_format=fmt, border=THIN_BORDER,
                    fill=fill)


def _write_label_row(ws, row, labels_values, max_col):
    """写入标签-值行（两列布局）"""
    for i, (label, val) in enumerate(labels_values):
        col = 1 + i * 4
        _apply_cell(ws, row, col, label,
                    font=LABEL_FONT, alignment=LEFT, border=THIN_BORDER)
        _apply_cell(ws, row, col + 1, val,
                    font=NORMAL_FONT, alignment=CENTER,
                    number_format=NUM_FMT_2D if isinstance(val, float) else None,
                    border=THIN_BORDER)
        # 合并4列中的后两列留空
        if col + 2 <= max_col:
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)

    ws.row_dimensions[row].height = 22


def export_to_excel(hv_system: HVSystem, file_path: str, mode: str = "detailed") -> str:
    """
    将负荷计算数据导出为格式化的Excel文件。

    参数：
        hv_system: 高压系统数据
        file_path: 保存路径
        mode: "summary"（一览表，仅汇总数据）
              "detailed"（详细表，包含完整设备明细）
    """
    wb = Workbook()
    # 删除默认sheet
    wb.remove(wb.active)

    if mode == "summary":
        # ── 一览表模式：只创建一个汇总工作表 ──
        _build_summary_overview_sheet(wb, hv_system)
    else:
        # ── 详细表模式：为每个子系统创建独立工作表 ──
        for sub in hv_system.subsystems:
            _build_subsystem_sheet(wb, sub)
        # 创建设备明细清单
        _build_equipment_detail_sheet(wb, hv_system)

    # ── 创建总10kV负荷计算表 ──
    _build_hv_summary_sheet(wb, hv_system)

    # 保存
    wb.save(file_path)
    return file_path


def _build_subsystem_sheet(wb, sub):
    """构建单个子系统的负荷计算Excel工作表"""
    sheet_name = sub.name[:31]  # Excel sheet name max 31 chars
    ws = wb.create_sheet(title=sheet_name)

    summary = calc_subsystem_summary(sub)
    max_col = 12  # 最大列数
    row = 1

    # 设置列宽
    col_widths = {1: 6, 2: 22, 3: 10, 4: 12, 5: 12, 6: 12,
                  7: 10, 8: 8, 9: 8, 10: 12, 11: 12, 12: 12}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── 标题行 ──
    _write_title_row(ws, row, f"{sub.name} — 0.4kV负荷计算表", max_col)
    row += 2

    # ── 一、设备组计算明细 ──
    _write_section_header(ws, row, "一、设备组计算明细", max_col)
    row += 1

    headers = ["序号", "设备组名称", "Pe(kW)", "有功小计∑Pc\n(kW)",
               "无功小计∑Qc\n(kvar)", "视在小计∑Sc\n(kVA)",
               "cosφ", "Kp", "Kq", "计算Pc\n(kW)", "计算Qc\n(kvar)", "计算Sc\n(kVA)"]
    _write_table_header(ws, row, headers)
    row += 1

    sfmt = [NUM_FMT_INT, None, NUM_FMT_1D, NUM_FMT_2D, NUM_FMT_2D, NUM_FMT_2D,
            NUM_FMT_PF, NUM_FMT_PF, NUM_FMT_PF, NUM_FMT_2D, NUM_FMT_2D, NUM_FMT_2D]

    total_pe = 0.0
    for idx, g in enumerate(sub.groups, 1):
        values = [
            idx, g.name,
            g.total_device_power,
            g.subtotal_pc, g.subtotal_qc, g.subtotal_sc,
            g.power_factor,
            g.kp, g.kq,
            g.computed_pc, g.computed_qc, g.computed_sc,
        ]
        _write_data_row(ws, row, values, formats=sfmt)
        total_pe += g.total_device_power
        row += 1

    # 合计行
    total_pc = sub.total_pc
    total_qc = sub.total_qc
    total_sc = sub.total_sc
    pf_before = summary['pf_before']
    sum_values = ["", "合计", total_pe,
                  sum(g.subtotal_pc for g in sub.groups),
                  sum(g.subtotal_qc for g in sub.groups),
                  sqrt(sum(g.subtotal_pc for g in sub.groups) ** 2 +
                       sum(g.subtotal_qc for g in sub.groups) ** 2),
                  "", "", "",
                  total_pc, total_qc, total_sc]
    _write_data_row(ws, row, sum_values, formats=sfmt, fill=SUBTOTAL_FILL)
    row += 2

    # ── 二、无功补偿 ──
    _write_section_header(ws, row, "二、无功补偿", max_col)
    row += 1

    comp_items = [
        ("补偿前功率因数 cosφ", pf_before, NUM_FMT_PF),
        ("目标功率因数", sub.target_power_factor, NUM_FMT_PF),
        ("需要补偿容量 (kvar)", summary['required_qc'], NUM_FMT_1D),
        ("实际补偿容量 (kvar)", sub.compensation_qc, NUM_FMT_INT),
        ("补偿后功率因数 cosφ", summary['pf_after'], NUM_FMT_PF),
        ("补偿后有功 Pc (kW)", summary['compensated_pc'], NUM_FMT_2D),
        ("补偿后无功 Qc (kvar)", summary['compensated_qc'], NUM_FMT_2D),
        ("补偿后视在 Sc (kVA)", summary['compensated_sc'], NUM_FMT_2D),
    ]
    for i, (label, val, nfmt) in enumerate(comp_items):
        _apply_cell(ws, row, 1, label,
                    font=LABEL_FONT, alignment=LEFT, border=THIN_BORDER)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        _apply_cell(ws, row, 3, val,
                    font=NORMAL_FONT, alignment=CENTER,
                    number_format=nfmt, border=THIN_BORDER)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        row += 1

    row += 1

    # ── 三、变压器选择 ──
    _write_section_header(ws, row, "三、变压器选择", max_col)
    row += 1

    tf_items = [
        ("单台变压器容量 (kVA)", sub.transformer_rating or 0, NUM_FMT_INT),
        ("变压器台数", sub.transformer_count or 0, NUM_FMT_INT),
        ("运行方式", sub.transformer_operation_mode, None),
        ("有效容量 (kVA)", sub.effective_transformer_capacity or 0, NUM_FMT_INT),
        ("变压器负载率", summary['load_rate'], NUM_FMT_PCT),
    ]
    for label, val, nfmt in tf_items:
        _apply_cell(ws, row, 1, label,
                    font=LABEL_FONT, alignment=LEFT, border=THIN_BORDER)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        _apply_cell(ws, row, 3, val,
                    font=NORMAL_FONT, alignment=CENTER,
                    number_format=nfmt, border=THIN_BORDER)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        row += 1

    row += 1

    # ── 四、10kV侧负荷 ──
    _write_section_header(ws, row, "四、10kV侧负荷（折算后）", max_col)
    row += 1

    hv_items = [
        ("变压器有功损耗 ΔP (kW)", summary['transformer_loss_p'], NUM_FMT_2D),
        ("变压器无功损耗 ΔQ (kvar)", summary['transformer_loss_q'], NUM_FMT_2D),
        ("高压侧有功 Pc (kW)", summary['hv_pc'], NUM_FMT_2D),
        ("高压侧无功 Qc (kvar)", summary['hv_qc'], NUM_FMT_2D),
        ("高压侧视在 Sc (kVA)", summary['hv_sc'], NUM_FMT_2D),
        ("高压侧功率因数 cosφ", summary['hv_pf'], NUM_FMT_PF),
    ]
    for label, val, nfmt in hv_items:
        _apply_cell(ws, row, 1, label,
                    font=LABEL_FONT, alignment=LEFT, border=THIN_BORDER)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        _apply_cell(ws, row, 3, val,
                    font=NORMAL_FONT, alignment=CENTER,
                    number_format=nfmt, border=THIN_BORDER)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        row += 1

    # 打印设置
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1


def _build_summary_overview_sheet(wb, hv_system):
    """构建一览表：所有子系统的汇总数据"""
    ws = wb.create_sheet(title="各系统汇总")
    max_col = 8
    row = 1

    col_widths = {1: 6, 2: 26, 3: 14, 4: 14, 5: 14, 6: 12, 7: 14, 8: 14}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    _write_title_row(ws, row, "各配电系统负荷汇总一览表", max_col)
    row += 2

    # 表头
    headers = ["序号", "子系统名称", "总有功Pc\n(kW)", "总无功Qc\n(kvar)",
               "总视在Sc\n(kVA)", "补偿前\ncosφ", "补偿容量\n(kvar)", "补偿后\ncosφ"]
    _write_table_header(ws, row, headers)
    row += 1

    total_pc = 0.0
    total_qc = 0.0
    for idx, sub in enumerate(hv_system.subsystems, 1):
        sm = calc_subsystem_summary(sub)
        values = [
            idx, sub.name,
            sm['pc'], sm['qc'], sm['sc'],
            sm['pf_before'],
            sub.compensation_qc,
            sm['pf_after'],
        ]
        fmts = [NUM_FMT_INT, None, NUM_FMT_2D, NUM_FMT_2D, NUM_FMT_2D,
                NUM_FMT_PF, NUM_FMT_INT, NUM_FMT_PF]
        _write_data_row(ws, row, values, formats=fmts)
        total_pc += sm['pc']
        total_qc += sm['qc']
        row += 1

    row += 1

    # 变压器与高压侧汇总（两列布局）
    _write_section_header(ws, row, "变压器与10kV侧汇总", max_col)
    row += 1

    for sub in hv_system.subsystems:
        sm = calc_subsystem_summary(sub)
        items = [
            (f"【{sub.name[:20]}】", "", None),
            ("变压器容量(kVA)", sm['transformer_capacity'], NUM_FMT_INT),
            ("负载率", sm['load_rate'], NUM_FMT_PCT),
            ("高压侧Pc(kW)", sm['hv_pc'], NUM_FMT_2D),
            ("高压侧Qc(kvar)", sm['hv_qc'], NUM_FMT_2D),
            ("高压侧Sc(kVA)", sm['hv_sc'], NUM_FMT_2D),
            ("高压侧cosφ", sm['hv_pf'], NUM_FMT_PF),
        ]
        for label, val, nfmt in items:
            _apply_cell(ws, row, 1, label,
                        font=LABEL_FONT, alignment=LEFT, border=THIN_BORDER)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            _apply_cell(ws, row, 3, val,
                        font=NORMAL_FONT if nfmt else BOLD_FONT,
                        alignment=CENTER,
                        number_format=nfmt, border=THIN_BORDER)
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
            row += 1
        row += 1

    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1


def _build_equipment_detail_sheet(wb, hv_system):
    """构建设备明细清单"""
    ws = wb.create_sheet(title="设备明细清单")
    max_col = 11
    row = 1

    col_widths = {1: 6, 2: 22, 3: 20, 4: 12, 5: 8, 6: 8, 7: 8, 8: 8, 9: 10,
                  10: 10, 11: 14}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    _write_title_row(ws, row, "全厂设备明细清单", max_col)
    row += 2

    headers = ["序号", "子系统", "设备名称", "Pe\n(kW)", "安装\n台数",
               "工作\n台数", "Kx", "cosφ", "Pc\n(kW)", "Qc\n(kvar)", "负荷等级"]
    _write_table_header(ws, row, headers)
    row += 1

    idx = 0
    for sub in hv_system.subsystems:
        for g in sub.groups:
            for eq in g.equipment_list:
                if eq.is_subtotal:
                    continue
                idx += 1
                values = [
                    idx, f"{sub.name}/{g.name}", eq.name,
                    eq.rated_power, eq.installed_count, eq.working_count,
                    eq.kx, eq.cos_phi, eq.pc, eq.qc,
                    eq.load_level,
                ]
                fmts = [NUM_FMT_INT, None, None, NUM_FMT_1D, NUM_FMT_INT,
                        NUM_FMT_INT, NUM_FMT_PF, NUM_FMT_PF, NUM_FMT_2D,
                        NUM_FMT_2D, None]
                _write_data_row(ws, row, values, formats=fmts)
                row += 1

    # 汇总
    total_pe = sum(e.rated_power * e.working_count
                   for sub in hv_system.subsystems
                   for g in sub.groups
                   for e in g.equipment_list if not e.is_subtotal)
    total_pc = sum(e.pc
                   for sub in hv_system.subsystems
                   for g in sub.groups
                   for e in g.equipment_list if not e.is_subtotal)
    total_qc = sum(e.qc
                   for sub in hv_system.subsystems
                   for g in sub.groups
                   for e in g.equipment_list if not e.is_subtotal)
    row += 1
    sum_values = ["", "合计", f"共{idx}台设备", total_pe, "", "", "", "", total_pc, total_qc, ""]
    sfmt = [NUM_FMT_INT, None, None, NUM_FMT_1D, None, None, None, None,
            NUM_FMT_2D, NUM_FMT_2D, None]
    _write_data_row(ws, row, sum_values, formats=sfmt, fill=SUBTOTAL_FILL)

    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1


def _build_hv_summary_sheet(wb, hv_system):
    """构建10kV总负荷计算表"""
    ws = wb.create_sheet(title="总10kV负荷计算")
    max_col = 8
    row = 1

    col_widths = {1: 6, 2: 22, 3: 14, 4: 14, 5: 14, 6: 12, 7: 16, 8: 12}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── 标题行 ──
    _write_title_row(ws, row, "全厂10kV总负荷计算表", max_col)
    row += 2

    # ── 各子系统10kV侧汇总 ──
    _write_section_header(ws, row, "一、各子系统10kV侧汇总", max_col)
    row += 1

    headers = ["序号", "子系统名称", "Pc(kW)", "Qc(kvar)", "Sc(kVA)",
               "cosφ", "有效容量(kVA)", "负载率(%)"]
    _write_table_header(ws, row, headers)
    row += 1

    total_pc = 0.0
    total_qc = 0.0
    total_sc = 0.0

    for idx, sub in enumerate(hv_system.subsystems, 1):
        sm = calc_subsystem_summary(sub)
        hv_pc = sm['hv_pc']
        hv_qc = sm['hv_qc']
        hv_sc = sm['hv_sc']
        total_pc += hv_pc
        total_qc += hv_qc

        values = [
            idx, sub.name,
            hv_pc, hv_qc, hv_sc,
            sm['hv_pf'],
            sub.effective_transformer_capacity or 0,
            sm['load_rate'] * 100,
        ]
        fmts = [NUM_FMT_INT, None, NUM_FMT_2D, NUM_FMT_2D, NUM_FMT_2D,
                NUM_FMT_PF, NUM_FMT_INT, NUM_FMT_1D]
        _write_data_row(ws, row, values, formats=fmts)
        row += 1

    total_sc = sqrt(total_pc ** 2 + total_qc ** 2)
    total_pf = total_pc / total_sc if total_sc > 0 else 0

    row += 1

    # ── 全厂总计 ──
    _write_section_header(ws, row, "二、全厂总计", max_col)
    row += 1

    total_items = [
        ("总有功功率 Pc (kW)", total_pc, NUM_FMT_2D),
        ("总无功功率 Qc (kvar)", total_qc, NUM_FMT_2D),
        ("总视在功率 Sc (kVA)", total_sc, NUM_FMT_2D),
        ("总功率因数 cosφ", total_pf, NUM_FMT_PF),
    ]
    ncols = 2
    for i, (label, val, nfmt) in enumerate(total_items, 1):
        col = (i - 1) % ncols * 4 + 1
        r = row + (i - 1) // ncols
        _apply_cell(ws, r, col, label,
                    font=LABEL_FONT, alignment=LEFT, border=THIN_BORDER)
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 1)
        _apply_cell(ws, r, col + 2, val,
                    font=BOLD_FONT, alignment=CENTER,
                    number_format=nfmt, border=THIN_BORDER)
        ws.merge_cells(start_row=r, start_column=col + 2, end_row=r, end_column=col + 3)

    # 打印设置
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
