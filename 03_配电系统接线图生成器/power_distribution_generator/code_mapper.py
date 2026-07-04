"""
代号映射模块

参照《氛围化编程指令书_配电系统图生成器.md》第2.3节。
从Excel"代号"工作表读取映射字典，将设备名称映射为电气代号，
用于自动生成完整的回路编号。

代号表包含4个子表：
- 工艺设备参照代号：设备名称→代号（如水泵→GP）
- 常用电气设备参照代号：设备名称→代号（如低压开关柜→AN）
- 给水工程代号：构筑物→代号
- 排水工程代号：构筑物→代号
"""

import openpyxl


class CodeMapper:
    """代号映射器

    从Excel"代号"工作表中读取4个子表的映射关系，
    根据回路用途自动匹配设备代号，生成完整回路编号。
    """

    # 内置默认代号映射（当Excel中无"代号"工作表时使用）
    DEFAULT_PROCESS_MAP = {
        "水泵": "GP",
        "污水泵": "GP",
        "给水泵": "GP",
        "排水泵": "GP",
        "循环泵": "GP",
        "污泥泵": "SP",
        "刮泥机": "HL",
        "搅拌机": "AG",
        "风机": "FN",
        "鼓风机": "FN",
        "曝气机": "AE",
        "起重机": "CR",
        "电动葫芦": "CR",
        "闸门": "GA",
        "启闭机": "GA",
        "格栅": "SC",
        "压滤机": "FP",
        "脱水机": "FP",
        "加药装置": "CD",
        "消毒装置": "DI",
        "照明": "AL",
        "插座": "AX",
        "空调": "AC",
        "通风机": "VF",
        "进线": "进线",
        "联络": "联络",
        "备用": "备用",
    }

    DEFAULT_ELECTRICAL_MAP = {
        "低压开关柜": "AN",
        "配电箱": "AL",
        "照明配电箱": "AL",
        "动力配电箱": "AP",
        "控制箱": "AC",
    }

    def __init__(self, excel_path: str = None):
        """初始化代号映射器

        Args:
            excel_path: Excel文件路径，包含"代号"工作表
        """
        self.process_map = dict(self.DEFAULT_PROCESS_MAP)
        self.electrical_map = dict(self.DEFAULT_ELECTRICAL_MAP)
        self.water_supply_map = {}
        self.water_drainage_map = {}

        if excel_path:
            self._load_from_excel(excel_path)

    def _load_from_excel(self, excel_path: str):
        """从Excel"代号"工作表加载映射

        参照《氛围化编程指令书_配电系统图生成器.md》第2.3节。

        Args:
            excel_path: Excel文件路径
        """
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
        except Exception:
            return  # 静默失败，使用默认映射

        if "代号" not in wb.sheetnames:
            wb.close()
            return

        ws = wb["代号"]
        current_section = None

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=3, values_only=False):
            values = [cell.value for cell in row]

            # 跳过空行
            if all(v is None for v in values):
                continue

            # 检测标题行
            if values[0] and isinstance(values[0], str):
                title = values[0].strip()
                if "工艺设备" in title or "参照代号" in title:
                    current_section = "process"
                    continue
                elif "常用电气设备" in title:
                    current_section = "electrical"
                    continue
                elif "给水工程" in title:
                    current_section = "water_supply"
                    continue
                elif "排水工程" in title:
                    current_section = "water_drainage"
                    continue

            # 根据当前区域读取映射
            if current_section and values[1] and values[2]:
                name = str(values[1]).strip()
                code = str(values[2]).strip()
                if current_section == "process":
                    self.process_map[name] = code
                elif current_section == "electrical":
                    self.electrical_map[name] = code
                elif current_section == "water_supply":
                    self.water_supply_map[name] = code
                elif current_section == "water_drainage":
                    self.water_drainage_map[name] = code

        wb.close()

    def get_code(self, circuit_usage: str) -> str:
        """根据回路用途获取设备代号

        先在工艺设备映射中查找，再尝试部分匹配。

        Args:
            circuit_usage: 回路用途（如"水泵"、"1#进线"）

        Returns:
            设备代号（如"GP"），未找到时返回空字符串
        """
        # 精确匹配
        if circuit_usage in self.process_map:
            return self.process_map[circuit_usage]

        # 部分匹配（截断数字前缀和后缀）
        for name, code in self.process_map.items():
            if name in circuit_usage or circuit_usage in name:
                return code

        # 特殊规则：进线
        if "进线" in circuit_usage:
            return "进线"

        return ""

    def generate_circuit_no(self, circuit_usage: str, panel_no: str,
                            existing_numbers: list) -> str:
        """生成完整回路编号

        参照《氛围化编程指令书_配电系统图生成器.md》第2.3节。

        格式：-{设备代号}{序号}
        如：-GP1、-AL1

        Args:
            circuit_usage: 回路用途
            panel_no: 开关柜编号（如"J-AN"）
            existing_numbers: 已存在的回路编号列表，用于生成序号

        Returns:
            完整回路编号
        """
        code = self.get_code(circuit_usage)
        if not code:
            return ""

        # 生成序号
        count = 1
        while True:
            candidate = f"-{code}{count}"
            if candidate not in existing_numbers:
                return candidate
            count += 1
