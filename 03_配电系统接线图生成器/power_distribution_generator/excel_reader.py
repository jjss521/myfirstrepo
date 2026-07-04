"""
Excel数据读取与解析模块

参照《氛围化编程指令书_配电系统图生成器.md》第2.1~2.2节。
使用 openpyxl 读取"低压配电系统"工作表，按"行=属性项，列=回路"的
横向展开格式解析为 PanelData 和 CircuitData 数据模型。

Excel布局说明：
- 每个开关柜占一组行（行2~18），B列=属性名，C~J列=回路1~8
- 多个开关柜竖向堆叠：行2~18为柜1，行20~36为柜2，以此类推
"""

import openpyxl
import logging
from models import PanelData, CircuitData
from calc_engine import calculate_circuit
from code_mapper import CodeMapper

logger = logging.getLogger(__name__)


# 单元格映射定义（第2.1节）
# (行号相对偏移, 属性名, 数据类型)
# 行2 = 相对偏移0（柜内首行）
CELL_MAP = {
    2: ("panel_no", str),           # 开关柜编号
    3: ("panel_type", str),         # 开关柜柜型
    4: ("panel_size", str),         # 开关柜尺寸
    5: ("unit_space", str),         # 单元空间（自动生成）
    6: ("circuit_usage", str),      # 回路用途
    7: ("pe_power", float),         # 设备功率Pe(kW) —— 手动输入
    8: ("ic_current", float),       # 计算电流Ic(A) —— 自动计算
    9: ("frame_current", int),      # 断路器壳架电流(A) —— 自动计算
    10: ("in_rated", int),          # 脱扣器额定电流In(A) —— 自动计算
    11: ("is1", float),             # 长延时整定Is1(A) —— 自动计算
    12: ("is2", float),             # 短延时整定Is2(A) —— 自动计算
    13: ("is3", float),             # 瞬动整定Is3(A) —— 自动计算
    14: ("ct_ratio", str),          # 电流互感器变比 —— 自动计算
    15: ("monitor", str),           # 电力监控信号 —— 自动计算
    16: ("cable_spec", str),        # 线缆型号规格
    17: ("cable_section", str),     # 线缆截面规格(预留)
    18: ("cable_no", str),          # 线缆编号
}

# 自动计算字段集合：Excel中即使有值也忽略，由 calc_engine 统一计算
# （第2.2节扩展：除手动输入外所有数据由引擎自动生成）
AUTO_CALC_FIELDS = {
    "unit_space",       # 行5
    "ic_current",       # 行8
    "frame_current",    # 行9
    "in_rated",         # 行10
    "is1",              # 行11
    "is2",              # 行12
    "is3",              # 行13
    "ct_ratio",         # 行14
    "monitor",          # 行15
}

# CABINET_ROW_COUNT = 17  # 每个柜占 18-2+1=17 行

# 回路列索引（C~J列，对应openpyxl的列索引3~10，但values列表是0-based索引2~9）
CIRCUIT_COLUMNS = list(range(2, 10))  # values[2] ~ values[9]


def read_panel_from_excel(file_path: str) -> list:
    """从Excel文件读取所有开关柜数据

    参照《氛围化编程指令书_配电系统图生成器.md》第2.1~2.2节。

    解析策略：
    1. 遍历所有行
    2. 当在行2位置检测到B列有"开关柜编号"文本且C~J列有柜号值时，识别为新柜开始
    3. 读取行2~18范围内的数据，形成完整柜数据
    4. 重复步骤2~3直到文件结束

    注意：由于 data_only=True 模式下无法直接通过公式判断行号，
    我们使用行号+属性名联合检测策略。

    Args:
        file_path: Excel文件路径

    Returns:
        PanelData列表
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)

    if "低压配电系统" not in wb.sheetnames:
        wb.close()
        return []

    ws = wb["低压配电系统"]
    mapper = CodeMapper(file_path)

    panels = []
    section_starts = _detect_cabinet_sections(ws)

    for start_row in section_starts:
        panel = _parse_cabinet_section(ws, start_row, mapper)
        if panel is not None:
            panels.append(panel)

    wb.close()
    return panels


def _detect_cabinet_sections(ws) -> list:
    """检测每个开关柜的起始行号

    检测方法：在行2（属性行为"开关柜编号"）检测C~J列是否有有效柜号值。

    Returns:
        每个柜起始行号的列表
    """
    sections = []
    row = 2
    max_row = ws.max_row or 200

    while row <= max_row:
        # 检查第row行B列是否为"开关柜编号"
        b_val = ws.cell(row=row, column=2).value
        if b_val is not None and str(b_val).strip() == "开关柜编号":
            # 检查C~J列是否有值
            has_data = False
            for col in range(3, 11):  # C(3) ~ J(10)
                val = ws.cell(row=row, column=col).value
                if val is not None and str(val).strip():
                    has_data = True
                    break
            if has_data:
                sections.append(row)

        row += 1

    return sections


def _parse_cabinet_section(ws, start_row: int, mapper: CodeMapper) -> PanelData:
    """解析一个开关柜的数据区间（start_row ~ start_row+16）

    Args:
        ws: 工作表对象
        start_row: 柜起始行（行2=柜1开始）
        mapper: 代号映射器

    Returns:
        PanelData 或 None（解析失败时）
    """
    panel = None

    # 存储柜级属性和回路级原始数据
    panel_attrs = {}
    circuit_raw_data = {}  # col_idx -> {field: value}

    for offset in range(17):  # 行2~18 = offset 0~16
        row_num = start_row + offset
        row_key = 2 + offset  # 映射到CELL_MAP中的行号

        if row_key not in CELL_MAP:
            continue

        attr_field, attr_type = CELL_MAP[row_key]

        # 读取B列属性值（整行共享的属性名）
        # C~J列 = 回路列（openpyxl列号3~10）
        for col in range(3, 11):  # C=3, D=4, ..., J=10
            cell_val = ws.cell(row=row_num, column=col).value

            # 空单元格跳过（第2.2节第4条）
            if cell_val is None:
                continue

            # 类型转换
            try:
                if attr_type == float:
                    processed = float(cell_val)
                elif attr_type == int:
                    processed = int(float(cell_val))
                else:
                    processed = str(cell_val).strip()
            except (ValueError, TypeError):
                processed = str(cell_val).strip()

            # openpyxl列号转0-based索引用于values
            col_idx = col - 1  # C=3 -> values[2]

            # 柜级别属性
            if row_key in (2, 3, 4):
                if row_key == 2:
                    # 开关柜编号：去除前导"="号（第2.2节第6条）
                    panel_no = str(cell_val).strip().lstrip("=")
                    if panel:
                        pass
                    else:
                        panel = PanelData(panel_no=panel_no)
                    panel_attrs["panel_no"] = panel_no
                elif row_key == 3:
                    panel_attrs["panel_type"] = processed
                    if panel:
                        panel.panel_type = processed
                elif row_key == 4:
                    panel_attrs["panel_size"] = processed
                    if panel:
                        panel.panel_size = processed
            else:
                # 回路级别属性
                # 跳过自动计算字段：这些由 calc_engine 统一计算，
                # Excel 中即使有值也不读取（确保引擎一致性）
                if attr_field in AUTO_CALC_FIELDS:
                    continue

                if col_idx not in circuit_raw_data:
                    circuit_raw_data[col_idx] = {
                        "column_index": col_idx,
                        "circuit_usage": "",
                    }
                circuit_raw_data[col_idx][attr_field] = processed

    if panel is None:
        return None

    # 补充柜属性
    for k, v in panel_attrs.items():
        setattr(panel, k, v)

    # 组装回路数据
    existing_numbers = []
    for col_idx in sorted(circuit_raw_data.keys()):
        data = circuit_raw_data[col_idx]

        # 跳过全空回路（无用途且无功率）
        if not data.get("circuit_usage", "") and not data.get("pe_power", 0):
            continue

        circuit = CircuitData(
            column_index=data.get("column_index", col_idx),
            circuit_usage=str(data.get("circuit_usage", "") or ""),
            pe_power=float(data.get("pe_power", 0) or 0),
            ic_current=float(data.get("ic_current", 0) or 0),
            frame_current=int(data.get("frame_current", 0) or 0),
            in_rated=int(data.get("in_rated", 0) or 0),
            is1=float(data.get("is1", 0) or 0),
            is2=float(data.get("is2", 0) or 0),
            is3=float(data.get("is3", 0) or 0),
            ct_ratio=str(data.get("ct_ratio", "") or ""),
            monitor=str(data.get("monitor", "") or ""),
            cable_spec=str(data.get("cable_spec", "") or ""),
            cable_section=str(data.get("cable_section", "") or ""),
            cable_no=str(data.get("cable_no", "") or ""),
            unit_space=str(data.get("unit_space", "") or ""),
        )

        # 执行电气计算（补充空缺计算值）
        circuit = calculate_circuit(circuit, panel.panel_type)

        # 生成回路编号
        circuit_no = mapper.generate_circuit_no(
            circuit.circuit_usage, panel.panel_no, existing_numbers
        )
        circuit.circuit_no = circuit_no
        if circuit_no:
            existing_numbers.append(circuit_no)

        panel.circuits.append(circuit)

    return panel


def print_panel_data(panels: list):
    """打印解析结果到控制台

    参照《氛围化编程指令书_配电系统图生成器.md》阶段1第5条。

    Args:
        panels: PanelData列表
    """
    for panel in panels:
        print(f"开关柜: {panel.panel_no} | 柜型: {panel.panel_type} "
              f"| 尺寸: {panel.panel_size}")
        for i, circuit in enumerate(panel.circuits):
            print(f"  回路{i+1}: {circuit.circuit_usage} | "
                  f"{circuit.pe_power}kW | {circuit.ic_current}A | "
                  f"壳架{circuit.frame_current}A | "
                  f"In{circuit.in_rated}A | "
                  f"CT {circuit.ct_ratio} | "
                  f"单元空间 {circuit.unit_space} | "
                  f"回路编号 {circuit.circuit_no}")
        print()


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        panels = read_panel_from_excel(sys.argv[1])
        print_panel_data(panels)
    else:
        print("用法: python excel_reader.py <Excel文件路径>")
