"""
示例Excel数据生成器

参照《氛围化编程指令书_配电系统图生成器.md》第2.1节单元格映射，
生成包含至少1个开关柜、8个回路的示例数据文件。

注意：仅供测试验证使用，实际生产中应使用设计院的原始Excel文件。
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def create_sample_excel(output_path: str):
    """生成示例配电系统设计Excel文件

    参照《氛围化编程指令书_配电系统图生成器.md》第2.1节结构：
    - "低压配电系统"工作表：行=属性，列（C~J）=回路
    - "代号"工作表：设备名→代号映射
    """
    wb = openpyxl.Workbook()

    # ========== 低压配电系统工作表 ==========
    ws = wb.active
    ws.title = "低压配电系统"

    # 样式定义
    header_font = Font(name="Microsoft YaHei", bold=True, size=10)
    data_font = Font(name="Microsoft YaHei", size=10)
    formula_font = Font(name="Microsoft YaHei", size=10, color="0066CC")
    source_font = Font(name="Microsoft YaHei", size=9, italic=True, color="888888")
    header_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
    auto_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 列宽
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    for col in range(3, 11):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions["L"].width = 16

    # ---- 数据（一组开关柜，含8个回路） ----
    # 每行格式：(行号, B列属性名, C~J列数据列表[8个值], L列来源说明)
    # None 表示空单元格
    # 注意：除 Pe(设备功率) 和 回路用途 外，其余计算字段全部留空，
    # 由 calc_engine 根据 Pe 自动计算所有电气参数。

    data_rows = [
        # 行2：开关柜编号
        (2, "开关柜编号",
         ["J-AN", "J-AN", "J-AN", "J-AN",
          "J-AN", "J-AN", "J-AN", "J-AN"],
         "手动输入"),

        # 行3：开关柜柜型
        (3, "开关柜柜型",
         ["抽屉柜"] * 8,
         "手动选择"),

        # 行4：开关柜尺寸
        (4, "开关柜尺寸",
         ["800*600*1800"] * 8,
         "手动选择"),

        # 行5：单元空间（自动计算，留空）
        (5, "单元空间",
         [None] * 8,
         "自动生成"),

        # 行6：回路用途
        (6, "回路用途",
         ["1#进线", "水泵", "风机", "排水泵",
          "刮泥机", "加药装置", "备用", "照明"],
         "手动选择"),

        # 行7：设备功率Pe(kW) —— 唯一手动输入的计算依据
        (7, "设备功率Pe(kW)",
         [100.0, 50.0, 37.0, 15.0, 5.5, 3.0, None, 8.0],
         "手动输入"),

        # 行8~15：全部自动计算，Excel中不填值
        (8, "计算电流Ic(A)",       [None] * 8, "自动生成"),
        (9, "断路器壳架电流(A)",   [None] * 8, "自动生成"),
        (10, "脱扣器额定电流In(A)", [None] * 8, "自动生成"),
        (11, "长延时整定Is1(A)",   [None] * 8, "自动生成"),
        (12, "短延时整定Is2(A)",   [None] * 8, "自动生成"),
        (13, "瞬动整定Is3(A)",     [None] * 8, "自动生成"),
        (14, "电流互感器变比",     [None] * 8, "自动生成"),
        (15, "电力监控信号",       [None] * 8, "自动生成"),

        # 行16：线缆型号（项目统一选择）
        (16, "线缆型号规格",
         ["YJV-0.6/1kV"] * 8,
         "项目统一"),

        # 行17~18：预留
        (17, "(预留)线缆截面规格", [None] * 8, "预留"),
        (18, "线缆编号",          [None] * 8, "预留"),
    ]

    # 写入数据
    for row_num, attr_name, circuit_values, source in data_rows:
        # B列 = 属性名
        cell_b = ws.cell(row=row_num, column=2, value=attr_name)
        cell_b.font = header_font
        cell_b.border = thin_border

        # C~J列 = 回路数据
        for i, val in enumerate(circuit_values):
            col = 3 + i  # C=3, D=4, ..., J=10
            cell = ws.cell(row=row_num, column=col)

            if val is not None:
                cell.value = val
            else:
                cell.value = None

            cell.font = formula_font if row_num in (8,) else data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # 自动生成字段用黄色背景标记
            if source in ("自动生成",):
                cell.fill = auto_fill

        # L列 = 来源说明
        cell_l = ws.cell(row=row_num, column=12, value=source)
        cell_l.font = source_font
        cell_l.border = thin_border
        cell_l.alignment = Alignment(horizontal="center", vertical="center")

    # ========== 代号工作表 ==========
    ws2 = wb.create_sheet("代号")

    # 列宽
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 14

    section_font = Font(name="Microsoft YaHei", bold=True, size=11, color="1565C0")

    # 表1：工艺设备参照代号
    ws2.cell(row=1, column=1, value="工艺设备参照代号").font = section_font
    ws2.cell(row=2, column=1, value="设备名称").font = header_font
    ws2.cell(row=2, column=2, value="代号").font = header_font
    ws2.cell(row=2, column=3, value="说明").font = header_font

    process_equip = [
        ("水泵", "GP", "给水泵、排水泵、循环泵等"),
        ("污水泵", "GP", ""),
        ("污泥泵", "SP", ""),
        ("刮泥机", "HL", ""),
        ("搅拌机", "AG", ""),
        ("风机", "FN", "鼓风机、通风机等"),
        ("曝气机", "AE", ""),
        ("起重机", "CR", "电动葫芦等"),
        ("闸门", "GA", "启闭机等"),
        ("格栅", "SC", ""),
        ("压滤机", "FP", "脱水机等"),
        ("加药装置", "CD", ""),
        ("消毒装置", "DI", ""),
        ("照明", "AL", ""),
        ("插座", "AX", ""),
        ("空调", "AC", ""),
    ]

    for i, (name, code, desc) in enumerate(process_equip):
        row = 3 + i
        ws2.cell(row=row, column=1, value=name).font = data_font
        ws2.cell(row=row, column=2, value=code).font = data_font
        ws2.cell(row=row, column=3, value=desc).font = data_font

    # 表2：常用电气设备参照代号
    offset = 3 + len(process_equip) + 2
    ws2.cell(row=offset, column=1, value="常用电气设备参照代号").font = section_font
    ws2.cell(row=offset + 1, column=1, value="设备名称").font = header_font
    ws2.cell(row=offset + 1, column=2, value="代号").font = header_font

    electrical_equip = [
        ("低压开关柜", "AN"),
        ("配电箱", "AL"),
        ("照明配电箱", "AL"),
        ("动力配电箱", "AP"),
        ("控制箱", "AC"),
    ]
    for i, (name, code) in enumerate(electrical_equip):
        row = offset + 2 + i
        ws2.cell(row=row, column=1, value=name).font = data_font
        ws2.cell(row=row, column=2, value=code).font = data_font

    # 保存
    wb.save(output_path)
    print(f"示例数据已生成: {output_path}")
    print("包含1个开关柜(J-AN)、8个回路")
    print("数据文件依据：《氛围化编程指令书_配电系统图生成器.md》第2.1节")


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        path = sys.argv[1]
    else:
        path = r"d:/qoderwork/power_distribution_generator/data/配电系统设计.xlsx"
    create_sample_excel(path)
