"""生成测试固件 Excel 文件

运行: python tests/generate_fixtures.py
"""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl

FIXTURES_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fixtures")


def create_sample_valid():
    """标准格式，20 条回路"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "回路清单"

    headers = [
        "回路编号", "回路名称", "负荷类型", "额定功率", "额定电流",
        "断路器型号", "断路器极数", "脱扣器电流", "CT变比",
        "电缆型号", "电缆规格", "变频器型号", "变频器功率", "备注",
    ]
    ws.append(headers)

    data = [
        ["L1", "1#水泵", "动力", 15.0, 28.5, "NSX100N", 3, 32.0, "50/5A", "YJV", "4x25+1x16", "", "", ""],
        ["L2", "2#水泵", "动力", 15.0, 28.5, "NSX100N", 3, 32.0, "50/5A", "YJV", "4x25+1x16", "", "", ""],
        ["L3", "3#水泵", "动力", 22.0, 42.0, "NSX160N", 3, 50.0, "100/5A", "YJV", "4x35+1x16", "", "", ""],
        ["L4", "1#风机", "动力", 11.0, 21.0, "NSX100N", 3, 25.0, "50/5A", "YJV", "4x16+1x10", "", "", ""],
        ["L5", "2#风机", "动力", 11.0, 21.0, "NSX100N", 3, 25.0, "50/5A", "YJV", "4x16+1x10", "", "", ""],
        ["L6", "变频器1", "变频", 30.0, 57.0, "NSX160N", 3, 63.0, "100/5A", "YJV", "4x50+1x25", "ATV320", 30.0, ""],
        ["L7", "变频器2", "变频", 22.0, 42.0, "NSX160N", 3, 50.0, "100/5A", "YJV", "4x35+1x16", "ATV320", 22.0, ""],
        ["L8", "1#空调", "空调", 18.5, 35.0, "NSX100N", 3, 40.0, "50/5A", "YJV", "4x25+1x16", "", "", ""],
        ["L9", "2#空调", "空调", 18.5, 35.0, "NSX100N", 3, 40.0, "50/5A", "YJV", "4x25+1x16", "", "", ""],
        ["L10", "3#空调", "空调", 7.5, 14.5, "NSX63N", 3, 16.0, "30/5A", "YJV", "4x10+1x6", "", "", ""],
        ["L11", "1层照明", "照明", 5.0, 9.5, "NSX63N", 1, 10.0, "20/5A", "BV", "3x4", "", "", ""],
        ["L12", "2层照明", "照明", 5.0, 9.5, "NSX63N", 1, 10.0, "20/5A", "BV", "3x4", "", "", ""],
        ["L13", "3层照明", "照明", 4.5, 8.5, "NSX63N", 1, 10.0, "20/5A", "BV", "3x4", "", "", ""],
        ["L14", "应急照明", "照明", 3.0, 5.7, "NSX63N", 1, 6.0, "10/5A", "BV", "3x2.5", "", "", "双电源"],
        ["L15", "1层插座", "插座", 3.0, 13.6, "NSX63N", 1, 16.0, "30/5A", "BV", "3x4", "", "", ""],
        ["L16", "2层插座", "插座", 3.0, 13.6, "NSX63N", 1, 16.0, "30/5A", "BV", "3x4", "", "", ""],
        ["L17", "备用1", "备用", 10.0, 19.0, "NSX63N", 3, 20.0, "30/5A", "YJV", "4x10+1x6", "", "", ""],
        ["L18", "备用2", "备用", 10.0, 19.0, "NSX63N", 3, 20.0, "30/5A", "YJV", "4x10+1x6", "", "", ""],
        ["L19", "电容补偿1", "电容补偿", 20.0, 28.9, "NSX100N", 3, 32.0, "50/5A", "YJV", "4x25+1x16", "", "", ""],
        ["L20", "电容补偿2", "电容补偿", 20.0, 28.9, "NSX100N", 3, 32.0, "50/5A", "YJV", "4x25+1x16", "", "", ""],
    ]
    for row in data:
        ws.append(row)

    path = os.path.join(FIXTURES_DIR, "sample_valid.xlsx")
    wb.save(path)
    print(f"已生成: {path}")


def create_sample_with_errors():
    """含错误数据（缺列、非法值、重复编号）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "回路清单"

    headers = [
        "回路编号", "回路名称", "负荷类型", "额定功率", "额定电流",
        "断路器型号", "断路器极数", "脱扣器电流", "CT变比",
        "电缆型号", "电缆规格", "变频器型号", "变频器功率", "备注",
    ]
    ws.append(headers)

    data = [
        # 正常行
        ["L1", "1#水泵", "动力", 15.0, 28.5, "NSX100N", 3, 32.0, "50/5A", "YJV", "4x25+1x16", "", "", ""],
        # 缺必填字段（回路名称为空）
        ["L2", "", "动力", 11.0, 21.0, "NSX100N", 3, 25.0, "50/5A", "YJV", "4x16+1x10", "", "", ""],
        # 非法负荷类型
        ["L3", "消防泵", "消防", 22.0, 42.0, "NSX160N", 3, 50.0, "100/5A", "YJV", "4x35+1x16", "", "", ""],
        # 重复编号
        ["L1", "2#水泵", "动力", 15.0, 28.5, "NSX100N", 3, 32.0, "50/5A", "YJV", "4x25+1x16", "", "", ""],
        # 数值越界（功率为负）
        ["L5", "3#水泵", "动力", -5.0, 28.5, "NSX100N", 3, 32.0, "50/5A", "YJV", "4x25+1x16", "", "", ""],
        # 极数非法（5P）
        ["L6", "4#水泵", "动力", 11.0, 21.0, "NSX100N", 5, 25.0, "50/5A", "YJV", "4x16+1x10", "", "", ""],
        # 变频器一致性（类型=变频但无变频器型号）
        ["L7", "变频测试", "变频", 30.0, 57.0, "NSX160N", 3, 63.0, "100/5A", "YJV", "4x50+1x25", "", "", ""],
        # 正常行
        ["L8", "1#风机", "动力", 11.0, 21.0, "NSX100N", 3, 25.0, "50/5A", "YJV", "4x16+1x10", "", "", ""],
    ]
    for row in data:
        ws.append(row)

    path = os.path.join(FIXTURES_DIR, "sample_with_errors.xlsx")
    wb.save(path)
    print(f"已生成: {path}")


def create_sample_edge_cases():
    """边界情况（变频/备用/极数变化）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "回路清单"

    headers = [
        "回路编号", "回路名称", "负荷类型", "额定功率", "额定电流",
        "断路器型号", "断路器极数", "脱扣器电流", "CT变比",
        "电缆型号", "电缆规格", "变频器型号", "变频器功率", "备注",
    ]
    ws.append(headers)

    data = [
        # 单相动力
        ["M1", "控制电源", "动力", 2.0, 8.7, "NSX63N", 1, 10.0, "20/5A", "BV", "3x4", "", "", ""],
        # 两相动力
        ["M2", "焊机", "动力", 8.0, 36.4, "NSX100N", 2, 40.0, "50/5A", "YJV", "3x16+1x10", "", "", ""],
        # 四相变频
        ["M3", "大型变频", "变频", 55.0, 104.5, "NSX250N", 4, 125.0, "150/5A", "YJV", "4x95+1x50", "ATV930", 55.0, ""],
        # 单相照明
        ["M4", "走廊灯", "照明", 1.5, 6.8, "NSX63N", 1, 10.0, "10/5A", "BV", "3x2.5", "", "", ""],
        # 三相照明
        ["M5", "景观照明", "照明", 10.0, 19.0, "NSX63N", 3, 20.0, "30/5A", "YJV", "4x10+1x6", "", "", ""],
        # 电容补偿
        ["M6", "无功补偿", "电容补偿", 30.0, 43.3, "NSX100N", 3, 50.0, "50/5A", "YJV", "4x35+1x16", "", "", ""],
        # 备用
        ["M7", "预留1", "备用", 0.1, 0.5, "NSX63N", 3, 10.0, "10/5A", "YJV", "4x6+1x4", "", "", "预留"],
    ]
    for row in data:
        ws.append(row)

    path = os.path.join(FIXTURES_DIR, "sample_edge_cases.xlsx")
    wb.save(path)
    print(f"已生成: {path}")


def create_sample_transposed():
    """转置格式测试文件（参数为行，回路为列）

    结构: A列=参数名, B列起每列=一个回路
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "低压配电系统"

    # 第一行: 表头
    ws.cell(row=1, column=1, value="参数")
    ws.cell(row=1, column=2, value="L1")
    ws.cell(row=1, column=3, value="L2")
    ws.cell(row=1, column=4, value="L3")
    ws.cell(row=1, column=5, value="L4")
    ws.cell(row=1, column=6, value="L5")

    # 参数行数据
    params = [
        ("开关柜代号",       "X-AN01",  "X-AN01",  "X-AN01",  "X-AN01",  "X-AN01"),
        ("开关柜尺寸(WxDH)mm", "600×600×2200", "", "", "", ""),
        ("单元空间",         "16E",     "16E",     "8E",      "8E",      "16E"),
        ("回路用途",         "1#水泵",  "变频器1",  "1层照明", "1#空调",  "电容补偿1"),
        ("设备功率Pe(kW)",   15.0,      30.0,      5.0,      18.5,     20.0),
        ("计算电流Ic(A)",    28.5,      57.0,      9.5,      35.0,     28.9),
        ("断路器壳架电流(A)", 100,       160,       63,       100,      100),
        ("脱扣器额定电流In(A)", 32,      63,        10,       40,       32),
        ("配电形式",         "MCC",     "MCC",     "PC",      "MCC",     "MCC"),
        ("运行方式",         "工频",    "变频",     "工频",    "工频",    "工频"),
        ("接触器",           "/",       "/",       "/",       "/",       "/"),
        ("热继电器",         "/",       "/",       "/",       "/",       "/"),
        ("变频器",           "/",       "200kW",   "/",       "/",       "/"),
        ("电流互感器变比",    "50",      "100",     "20",      "50",      "50"),
        ("电力监控信号",     "三相电流", "三相电流,有功电度", "照明控制", "温度", "无功"),
        ("线缆型号规格",     "YJV-0.6/1kV 4x25+1x16", "YJV-0.6/1kV 4x50+1x25", "BV 3x4", "YJV-0.6/1kV 4x25+1x16", "YJV-0.6/1kV 4x25+1x16"),
        ("线缆编号",         "L1",      "L2",      "L3",      "L4",      "L5"),
    ]

    for i, row_data in enumerate(params, start=2):
        for j, val in enumerate(row_data):
            ws.cell(row=i, column=j+1, value=val)

    path = os.path.join(FIXTURES_DIR, "sample_transposed.xlsx")
    wb.save(path)
    print(f"已生成: {path}")


def create_sample_transposed_errors():
    """转置格式含错误数据"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "低压配电系统"

    ws.cell(row=1, column=1, value="参数")
    ws.cell(row=1, column=2, value="E1")
    ws.cell(row=1, column=3, value="E1")  # 重复编号
    ws.cell(row=1, column=4, value="E3")

    params = [
        ("回路用途",         "", "2#水泵", "3#风机"),        # E1缺名称
        ("设备功率Pe(kW)",   15.0, 15.0, -5.0),              # E3负功率
        ("计算电流Ic(A)",    28.5, 28.5, 28.5),
        ("脱扣器额定电流In(A)", 32, 32, 32),
        ("电流互感器变比",    "50", "50", "50"),
        ("线缆型号规格",     "YJV 4x25+1x16", "YJV 4x25+1x16", "YJV 4x25+1x16"),
        ("运行方式",         "工频", "工频", "工频"),
        ("配电形式",         "MCC", "MCC", "MCC"),
    ]

    for i, row_data in enumerate(params, start=2):
        for j, val in enumerate(row_data):
            ws.cell(row=i, column=j+1, value=val)

    path = os.path.join(FIXTURES_DIR, "sample_transposed_errors.xlsx")
    wb.save(path)
    print(f"已生成: {path}")


if __name__ == "__main__":
    os.makedirs(FIXTURES_DIR, exist_ok=True)
    create_sample_valid()
    create_sample_with_errors()
    create_sample_edge_cases()
    create_sample_transposed()
    create_sample_transposed_errors()
    print("所有测试固件生成完成")
