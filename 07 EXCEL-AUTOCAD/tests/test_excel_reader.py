"""Excel 读取器单元测试"""
import os
import sys
import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.data_model import ExcelConfig, ExcelFormat, LoadType
from src.excel_reader import read_and_validate, _parse_float, _parse_int, _detect_format, _split_cable_full
from src.errors import ExcelReadError
from tests.conftest import (
    sample_valid_path, sample_errors_path, sample_edge_path,
    sample_transposed_path, sample_transposed_errors_path,
    default_config_path,
)
from src.config_loader import load_config


@pytest.fixture
def excel_cfg():
    cfg = load_config(default_config_path())
    return cfg.excel


class TestReadAndValidate:
    """read_and_validate 主流程测试"""

    def test_normal_read(self, excel_cfg):
        """正常读取标准格式 Excel"""
        records, errors = read_and_validate(sample_valid_path(), excel_cfg)
        assert len(records) == 20
        assert len(errors) == 0

    def test_record_fields(self, excel_cfg):
        """验证记录字段正确性"""
        records, _ = read_and_validate(sample_valid_path(), excel_cfg)
        r = records[0]
        assert r.circuit_id == "L1"
        assert r.circuit_name == "1#水泵"
        assert r.load_type == LoadType.POWER
        assert r.rated_power_kw == 15.0
        assert r.rated_current_a == 28.5
        assert r.breaker_model == "NSX100N"
        assert r.breaker_poles == 3
        assert r.ct_ratio == "50/5A"

    def test_vfd_record(self, excel_cfg):
        """验证变频器回路记录"""
        records, _ = read_and_validate(sample_valid_path(), excel_cfg)
        vfd_records = [r for r in records if r.load_type == LoadType.VFD]
        assert len(vfd_records) == 2
        assert vfd_records[0].vfd_model == "ATV320"
        assert vfd_records[0].vfd_power_kw == 30.0

    def test_error_data(self, excel_cfg):
        """含错误数据的 Excel"""
        records, errors = read_and_validate(sample_errors_path(), excel_cfg)
        # L1(正常) + L8(正常) = 2 条有效
        assert len(records) == 2
        # L2(空名称) + L3(非法类型) + L1重复 + L5(负功率) + L6(5P) + L7(变频无VFD) = 6 条错误
        assert len(errors) == 6

    def test_file_not_found(self, excel_cfg):
        """文件不存在"""
        with pytest.raises(ExcelReadError, match="不存在"):
            read_and_validate("nonexistent.xlsx", excel_cfg)

    def test_sheet_not_found(self, excel_cfg):
        """Sheet 未找到（所有回退名也不匹配）"""
        cfg = ExcelConfig(sheet_name="完全不存在_abc123")
        cfg.column_aliases = excel_cfg.column_aliases
        with pytest.raises(ExcelReadError, match="未找到"):
            read_and_validate(sample_valid_path(), cfg)

    def test_sheet_fallback(self, excel_cfg):
        """Sheet 名不匹配时自动回退到已知 Sheet"""
        cfg = ExcelConfig(sheet_name="低压配电系统")
        cfg.column_aliases = excel_cfg.column_aliases
        # sample_valid.xlsx 的 Sheet 是 "回路清单"，"低压配电系统" 应自动回退到 "回路清单"
        records, _ = read_and_validate(sample_valid_path(), cfg)
        assert len(records) > 0

    def test_edge_cases(self, excel_cfg):
        """边界情况（极数变化、变频、备用等）"""
        records, errors = read_and_validate(
            os.path.join(os.path.dirname(sample_valid_path()), "sample_edge_cases.xlsx"),
            excel_cfg,
        )
        assert len(records) == 7
        # 单相动力
        assert records[0].breaker_poles == 1
        # 四相变频
        assert records[2].load_type == LoadType.VFD
        assert records[2].breaker_poles == 4


class TestParseFunctions:
    """数值解析辅助函数测试"""

    def test_parse_float_normal(self):
        assert _parse_float("100") == 100.0
        assert _parse_float("28.5") == 28.5

    def test_parse_float_with_unit(self):
        assert _parse_float("100A") == 100.0
        assert _parse_float("15.5kW") == 15.5

    def test_parse_float_invalid(self):
        with pytest.raises(ValueError):
            _parse_float("abc")

    def test_parse_int_normal(self):
        assert _parse_int("3") == 3
        assert _parse_int("1P") == 1

    def test_parse_int_invalid(self):
        with pytest.raises(ValueError):
            _parse_int("abc")


class TestColumnAlias:
    """列名模糊匹配测试"""

    def test_alias_match(self, excel_cfg):
        """别名匹配：使用不同列名的 Excel"""
        # 使用标准文件，应该正常读取
        records, _ = read_and_validate(sample_valid_path(), excel_cfg)
        assert len(records) > 0


class TestTransposedFormat:
    """转置格式 Excel 读取测试"""

    def test_detect_transposed_format(self, excel_cfg):
        """自动检测转置格式"""
        import openpyxl
        wb = openpyxl.load_workbook(sample_transposed_path(), data_only=True)
        ws = wb[wb.sheetnames[0]]
        fmt = _detect_format(ws, excel_cfg)
        assert fmt == ExcelFormat.TRANSPOSED
        wb.close()

    def test_detect_standard_format(self, excel_cfg):
        """自动检测标准格式"""
        import openpyxl
        wb = openpyxl.load_workbook(sample_valid_path(), data_only=True)
        ws = wb[wb.sheetnames[0]]
        fmt = _detect_format(ws, excel_cfg)
        assert fmt == ExcelFormat.STANDARD
        wb.close()

    def test_read_transposed_valid(self, excel_cfg):
        """正常读取转置格式 Excel"""
        records, errors = read_and_validate(sample_transposed_path(), excel_cfg)
        assert len(records) == 5
        assert len(errors) == 0

    def test_transposed_record_fields(self, excel_cfg):
        """验证转置格式记录字段正确性"""
        records, _ = read_and_validate(sample_transposed_path(), excel_cfg)
        r = records[0]
        assert r.circuit_id == "L1"
        assert r.circuit_name == "1#水泵"
        assert r.load_type == LoadType.POWER
        assert r.rated_power_kw == 15.0
        assert r.rated_current_a == 28.5
        assert r.breaker_trip_current_a == 32.0
        assert r.cabinet_code == "X-AN01"

    def test_transposed_vfd_record(self, excel_cfg):
        """验证转置格式变频器回路"""
        records, _ = read_and_validate(sample_transposed_path(), excel_cfg)
        vfd_records = [r for r in records if r.load_type == LoadType.VFD]
        assert len(vfd_records) >= 1
        assert vfd_records[0].circuit_name == "变频器1"

    def test_transposed_cable_split(self, excel_cfg):
        """验证转置格式线缆型号规格拆分"""
        records, _ = read_and_validate(sample_transposed_path(), excel_cfg)
        r = records[0]
        assert "YJV" in r.cable_type
        assert "4x25" in r.cable_section or "25" in r.cable_section

    def test_transposed_new_fields(self, excel_cfg):
        """验证转置格式新增字段"""
        records, _ = read_and_validate(sample_transposed_path(), excel_cfg)
        r = records[0]
        assert r.distribution_type == "MCC"
        assert r.operation_mode == "工频"
        assert r.breaker_frame_current_a == 100.0
        assert r.unit_space == "16E"

    def test_transposed_ct_ratio_format(self, excel_cfg):
        """验证转置格式CT变比自动补全"""
        records, _ = read_and_validate(sample_transposed_path(), excel_cfg)
        # 纯数字"50"应被补全为"50/5A"
        assert records[0].ct_ratio == "50/5A"

    def test_transposed_with_errors(self, excel_cfg):
        """转置格式含错误数据"""
        records, errors = read_and_validate(sample_transposed_errors_path(), excel_cfg)
        # E1(空名称应跳过) + E1重复(跳过) + E3(负功率跳过) = 最多0-1条有效
        assert len(errors) >= 1

    def test_standard_still_works(self, excel_cfg):
        """确保标准格式仍然正常工作（回归测试）"""
        records, errors = read_and_validate(sample_valid_path(), excel_cfg)
        assert len(records) == 20
        assert len(errors) == 0

    def test_edge_cases_still_works(self, excel_cfg):
        """确保边界用例仍然正常工作（回归测试）"""
        records, errors = read_and_validate(sample_edge_path(), excel_cfg)
        assert len(records) == 7


class TestCableSplit:
    """线缆型号规格拆分测试"""

    def test_split_yjv_with_space(self):
        t, s = _split_cable_full("YJV-0.6/1kV 4x185+1x95")
        assert t == "YJV-0.6/1kV"
        assert s == "4x185+1x95"

    def test_split_bv_with_space(self):
        t, s = _split_cable_full("BV 3x4")
        assert t == "BV"
        assert s == "3x4"

    def test_split_empty(self):
        t, s = _split_cable_full("")
        assert t == ""
        assert s == ""

    def test_split_slash(self):
        t, s = _split_cable_full("/")
        assert t == ""
        assert s == ""
