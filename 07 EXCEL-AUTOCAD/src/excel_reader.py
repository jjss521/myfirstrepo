"""PDSG Excel 读取器

功能:
- 打开 .xlsx 工作簿
- 自动检测 Excel 格式（标准/转置）
- 标准格式: 每行一个回路，列为参数，三级模糊列名匹配
- 转置格式: 每列一个回路，行为参数（A列为参数名）
- 逐行/逐列读取 + 类型转换 + 校验
- 输出 List[CircuitRecord] + List[ErrorRecord]
"""
import logging
import os
import re
from typing import Dict, List, Optional, Tuple

import openpyxl

from .data_model import (
    CircuitRecord,
    ErrorRecord,
    ExcelConfig,
    ExcelFormat,
    LoadType,
    OPERATION_MODE_TO_LOAD_TYPE,
    RawRow,
)
from .errors import ExcelReadError

logger = logging.getLogger(__name__)

# 必填字段内部名（标准格式）
REQUIRED_FIELDS = [
    "circuit_id",
    "circuit_name",
    "load_type",
    "rated_power_kw",
    "rated_current_a",
    "breaker_model",
    "breaker_poles",
    "breaker_trip_current_a",
    "ct_ratio",
    "cable_type",
    "cable_section",
]

# 转置格式必填参数行名 → 内部字段名映射（默认映射，可被 transposed_column_aliases 覆盖）
DEFAULT_TRANSPOSED_MAPPING = {
    "回路用途": "circuit_name",
    "设备功率pe(kw)": "rated_power_kw",
    "计算电流ic(a)": "rated_current_a",
    "断路器壳架电流(a)": "breaker_frame_current_a",
    "脱扣器额定电流in(a)": "breaker_trip_current_a",
    "电流互感器变比": "ct_ratio",
    "线缆型号规格": "cable_full",
    "线缆编号": "cable_number",
    "运行方式": "operation_mode",
    "配电形式": "distribution_type",
    "接触器": "contactor",
    "热继电器": "thermal_relay",
    "变频器": "vfd_model",
    "电力监控信号": "power_monitoring",
    "开关柜代号": "cabinet_code",
    "开关柜尺寸(wxdh)mm": "cabinet_size",
    "单元空间": "unit_space",
}


def read_and_validate(
    path: str, cfg: ExcelConfig
) -> Tuple[List[CircuitRecord], List[ErrorRecord]]:
    """读取 Excel 并校验（自动检测格式）

    Args:
        path: Excel 文件路径
        cfg: Excel 配置

    Returns:
        (有效回路列表, 错误记录列表)

    Raises:
        ExcelReadError: 文件不存在/Sheet未找到/列匹配率不足
    """
    if not os.path.isfile(path):
        raise ExcelReadError(f"Excel 文件不存在: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = _find_sheet(wb, cfg.sheet_name)

    # 自动检测格式
    fmt = _detect_format(ws, cfg)
    logger.info("检测到 Excel 格式: %s", fmt.value)

    if fmt == ExcelFormat.TRANSPOSED:
        records, errors = _read_transposed(ws, cfg)
    else:
        records, errors = _read_standard(ws, cfg)

    wb.close()

    logger.info("校验完成: 有效 %d / 跳过 %d", len(records), len(errors))
    return records, errors


# ============================================================
# Sheet 查找
# ============================================================

def _find_sheet(wb: openpyxl.Workbook, sheet_name: str):
    """查找 Sheet，未找到时尝试常见别名，最后列出可用 Sheet

    仅当配置的名称是已知标准名时才启用回退：
    '低压配电系统' ↔ '回路清单' 互为回退。
    其他自定义名称不做回退。
    """
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]

    # 仅对已知标准名启用回退
    known_pairs = {
        "低压配电系统": "回路清单",
        "回路清单": "低压配电系统",
    }
    if sheet_name in known_pairs:
        fallback = known_pairs[sheet_name]
        if fallback in wb.sheetnames:
            logger.info("Sheet \"%s\" 未找到，自动使用备选 \"%s\"", sheet_name, fallback)
            return wb[fallback]

    available = ", ".join(wb.sheetnames)
    raise ExcelReadError(
        f"Sheet \"{sheet_name}\" 未找到。可用 Sheet: {available}"
    )


# ============================================================
# 格式检测
# ============================================================

def _detect_format(ws, cfg: ExcelConfig) -> ExcelFormat:
    """检测 Excel 文件格式

    通过检查 A1 单元格值判断：
    - "参数" → 转置格式（参数为行，回路为列）
    - 其他 → 标准格式（回路为行，列为参数）
    """
    if not cfg.format_auto_detect:
        return ExcelFormat.STANDARD

    a1_value = ws.cell(row=1, column=1).value
    if a1_value is not None and str(a1_value).strip() == "参数":
        return ExcelFormat.TRANSPOSED

    return ExcelFormat.STANDARD


# ============================================================
# 标准格式读取（原有逻辑）
# ============================================================

def _read_standard(
    ws, cfg: ExcelConfig
) -> Tuple[List[CircuitRecord], List[ErrorRecord]]:
    """读取标准格式 Excel（每行一个回路）"""
    header_row = cfg.header_row
    header_map = _read_header(ws, header_row)
    col_map = _match_columns(header_map, cfg.column_aliases)
    _check_match_ratio(col_map, cfg.min_match_ratio)

    logger.info(
        "标准格式: Sheet \"%s\" 共 %d 行数据（从第 %d 行起）",
        cfg.sheet_name,
        ws.max_row - cfg.data_start_row + 1,
        cfg.data_start_row,
    )

    records: List[CircuitRecord] = []
    errors: List[ErrorRecord] = []
    seen_ids: set = set()

    for row_num in range(cfg.data_start_row, ws.max_row + 1):
        raw = _read_row(ws, row_num, col_map)
        if raw is None:
            continue

        record, error = _validate_row(raw, seen_ids, cfg)
        if error:
            errors.append(error)
        if record:
            seen_ids.add(record.circuit_id)
            records.append(record)

    return records, errors


def _read_header(ws, header_row: int) -> Dict[int, str]:
    """读取表头行，返回 {列号: 列名}"""
    header = {}
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        if cell.value is not None:
            header[col] = str(cell.value).strip()
    return header


def _match_columns(
    header_map: Dict[int, str],
    aliases: Dict[str, str],
) -> Dict[str, int]:
    """列名三级模糊匹配

    优先级: 精确匹配 -> 别名匹配 -> 包含匹配
    返回 {内部字段名: 列号}
    """
    col_map: Dict[str, int] = {}

    alias_lower = {k.lower().strip(): v for k, v in aliases.items()}

    for col_num, col_name in header_map.items():
        col_name_stripped = col_name.strip()
        col_name_lower = col_name_stripped.lower()

        if col_name_lower in REQUIRED_FIELDS:
            col_map[col_name_lower] = col_num
            continue

        if col_name_lower in alias_lower:
            field_name = alias_lower[col_name_lower]
            col_map[field_name] = col_num
            continue

        matched = False
        for alias_key, field_name in alias_lower.items():
            if alias_key in col_name_lower or col_name_lower in alias_key:
                if field_name not in col_map:
                    col_map[field_name] = col_num
                    matched = True
                    break
        if not matched:
            logger.debug("未识别列: \"%s\" (列 %d)", col_name_stripped, col_num)

    return col_map


def _check_match_ratio(col_map: Dict[str, int], min_ratio: float) -> None:
    """检查必填列匹配率"""
    matched = sum(1 for f in REQUIRED_FIELDS if f in col_map)
    ratio = matched / len(REQUIRED_FIELDS) if REQUIRED_FIELDS else 1.0

    if ratio < min_ratio:
        identified = [f for f in REQUIRED_FIELDS if f in col_map]
        missing = [f for f in REQUIRED_FIELDS if f not in col_map]
        raise ExcelReadError(
            f"必填列匹配率不足 ({ratio:.0%} < {min_ratio:.0%})。"
            f"已识别: {identified}；缺失: {missing}"
        )
    logger.debug("必填列匹配率: %d/%d (%.0f%%)", matched, len(REQUIRED_FIELDS), ratio * 100)


def _read_row(ws, row_num: int, col_map: Dict[str, int]) -> Optional[RawRow]:
    """读取单行数据，空行返回 None"""
    values = {}
    for field_name, col_num in col_map.items():
        cell = ws.cell(row=row_num, column=col_num)
        if cell.value is not None:
            values[field_name] = str(cell.value).strip()

    if not values or all(v == "" or v == "None" for v in values.values()):
        return None

    return RawRow(row_number=row_num, values=values)


# ============================================================
# 转置格式读取（v1.1 新增）
# ============================================================

def _read_transposed(
    ws, cfg: ExcelConfig
) -> Tuple[List[CircuitRecord], List[ErrorRecord]]:
    """读取转置格式 Excel（A列为参数名，B列起每列为一个回路）

    结构示例:
        A列: 参数名        B列: L1    C列: L2    ...
        Row 1: 参数         L1        L2
        Row 2: 开关柜代号   X-AN01    ...
        Row 3: 回路用途     1#进线    1#进线
        ...
    """
    # 构建参数行映射: {参数名(小写): 行号}
    param_row_map: Dict[str, int] = {}
    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=1).value
        if cell_val is not None:
            param_row_map[str(cell_val).strip().lower()] = row

    if not param_row_map:
        raise ExcelReadError("转置格式: A列无有效参数名")

    # 合并用户自定义别名到默认映射
    alias_map = dict(DEFAULT_TRANSPOSED_MAPPING)
    if cfg.transposed_column_aliases:
        for alias_key, field_name in cfg.transposed_column_aliases.items():
            alias_map[alias_key.strip().lower()] = field_name

    # 解析参数行名 → 内部字段名
    field_row_map: Dict[str, int] = {}
    for param_name_lower, row_num in param_row_map.items():
        if param_name_lower in alias_map:
            field_name = alias_map[param_name_lower]
            field_row_map[field_name] = row_num
        else:
            # 尝试包含匹配
            for alias_key, field_name in alias_map.items():
                if alias_key in param_name_lower or param_name_lower in alias_key:
                    if field_name not in field_row_map:
                        field_row_map[field_name] = row_num
                        break

    logger.info(
        "转置格式: Sheet \"%s\" 共 %d 个参数行, %d 列回路",
        cfg.sheet_name,
        len(param_row_map),
        ws.max_column - 1,
    )
    logger.debug("转置字段映射: %s", list(field_row_map.keys()))

    # 检查基本字段是否可映射
    _check_transposed_fields(field_row_map, cfg.min_match_ratio)

    # 逐列读取回路（B列起）
    records: List[CircuitRecord] = []
    errors: List[ErrorRecord] = []
    seen_ids: set = set()

    for col in range(2, ws.max_column + 1):
        raw = _read_transposed_column(ws, col, field_row_map)
        if raw is None:
            continue

        record, error = _validate_transposed_row(raw, seen_ids, cfg)
        if error:
            errors.append(error)
        if record:
            seen_ids.add(record.circuit_id)
            records.append(record)

    return records, errors


def _check_transposed_fields(
    field_row_map: Dict[str, int],
    min_ratio: float,
) -> None:
    """检查转置格式的基本字段映射覆盖率"""
    # 转置格式的"必填字段"（比标准格式宽松）
    transposed_required = [
        "circuit_name", "rated_power_kw", "rated_current_a",
        "breaker_trip_current_a", "ct_ratio", "cable_full",
    ]
    matched = sum(1 for f in transposed_required if f in field_row_map)
    ratio = matched / len(transposed_required) if transposed_required else 1.0

    if ratio < min_ratio:
        identified = [f for f in transposed_required if f in field_row_map]
        missing = [f for f in transposed_required if f not in field_row_map]
        raise ExcelReadError(
            f"转置格式: 必填参数行匹配率不足 ({ratio:.0%} < {min_ratio:.0%})。"
            f"已识别: {identified}；缺失: {missing}"
        )


def _read_transposed_column(
    ws, col: int, field_row_map: Dict[str, int]
) -> Optional[RawRow]:
    """读取转置格式中的单列数据（一个回路）

    Args:
        ws: 工作表
        col: 列号（从2开始）
        field_row_map: {内部字段名: 行号}

    Returns:
        RawRow 或 None（空列）
    """
    values = {}
    for field_name, row_num in field_row_map.items():
        cell = ws.cell(row=row_num, column=col)
        if cell.value is not None:
            values[field_name] = str(cell.value).strip()

    # 读取回路编号（来自第一行的列标题）
    header_val = ws.cell(row=1, column=col).value
    if header_val is not None:
        values["circuit_id"] = str(header_val).strip()

    # 空列检测
    if not values or all(v == "" or v == "None" for v in values.values()):
        return None

    # 如果没有从标题获取到 circuit_id，使用列号生成
    if "circuit_id" not in values or values["circuit_id"] == "":
        if "cable_number" in values and values["cable_number"]:
            values["circuit_id"] = values["cable_number"]
        else:
            values["circuit_id"] = f"COL_{col}"

    # 行号用第2行（首个参数行）作为标识
    row_num = field_row_map.get("circuit_name", 2)
    return RawRow(row_number=col, values=values)


def _validate_transposed_row(
    raw: RawRow, seen_ids: set, cfg: ExcelConfig
) -> Tuple[Optional[CircuitRecord], Optional[ErrorRecord]]:
    """校验转置格式的单列数据"""
    v = raw.values

    def err(error_type: str, msg: str):
        return ErrorRecord(
            row_number=raw.row_number,
            circuit_id=v.get("circuit_id", ""),
            error_type=error_type,
            error_message=msg,
            raw_values=dict(v),
        )

    # 回路编号
    cid = v.get("circuit_id", "")
    if not cid or cid == "None":
        return None, err("缺回路编号", f"列 {raw.row_number}: 回路编号为空")

    # 回路编号唯一性
    if cid in seen_ids:
        return None, err("编号重复", f"列 {raw.row_number}: 回路编号 \"{cid}\" 重复")

    # 回路名称
    circuit_name = v.get("circuit_name", "")
    if not circuit_name or circuit_name == "None":
        return None, err("缺回路名称", f"列 {raw.row_number}: 回路用途为空")

    # 负荷类型推断：优先从运行方式推断，其次从配电形式
    load_type = _infer_load_type_from_transposed(v)
    if load_type is None:
        return None, err(
            "无法推断负荷类型",
            f"列 {raw.row_number}: 无法从运行方式/配电形式推断负荷类型",
        )

    # 数值字段
    try:
        rated_power = _parse_float(v.get("rated_power_kw", "0"))
    except ValueError:
        return None, err("数值转换失败", f"列 {raw.row_number}: 设备功率无法转换")

    try:
        rated_current = _parse_float(v.get("rated_current_a", "0"))
    except ValueError:
        return None, err("数值转换失败", f"列 {raw.row_number}: 计算电流无法转换")

    # 脱扣器额定电流
    breaker_trip = 0.0
    if "breaker_trip_current_a" in v and v["breaker_trip_current_a"]:
        try:
            breaker_trip = _parse_float(v["breaker_trip_current_a"])
        except ValueError:
            breaker_trip = 0.0

    # 断路器壳架电流（可选）
    breaker_frame = None
    if "breaker_frame_current_a" in v and v["breaker_frame_current_a"]:
        try:
            breaker_frame = _parse_float(v["breaker_frame_current_a"])
        except ValueError:
            breaker_frame = None

    # 断路器极数推断：从脱扣器电流和壳架电流推断
    breaker_poles = _infer_poles_from_transposed(v, rated_current)

    # 数值范围
    if rated_power <= 0:
        return None, err("数值越界", f"列 {raw.row_number}: 设备功率须>0，实际={rated_power}")
    if rated_current <= 0:
        return None, err("数值越界", f"列 {raw.row_number}: 计算电流须>0，实际={rated_current}")

    # 断路器型号（转置格式通常不含此列，使用配置默认值）
    breaker_model = v.get("breaker_model", "") or cfg.default_breaker_model or ""

    # CT 变比
    ct_ratio = v.get("ct_ratio", "")
    if ct_ratio and ct_ratio != "/":
        # 纯数字时补上格式（如 "400" → "400/5A"）
        try:
            ct_val = int(ct_ratio)
            ct_ratio = f"{ct_val}/5A"
        except ValueError:
            pass  # 已经是完整格式

    # 线缆型号规格拆分
    cable_type, cable_section = _split_cable_full(v.get("cable_full", ""))

    # 变频器相关
    vfd_model = v.get("vfd_model", None)
    vfd_power = None
    if vfd_model and vfd_model != "/" and vfd_model != "":
        # 解析变频器功率（如 "200kW" → 200.0）
        try:
            vfd_power = _parse_float(vfd_model)
        except ValueError:
            vfd_power = None
        # 如果变频器字段含型号信息，保留原始值
    else:
        vfd_model = None

    # 运行方式判断是否为变频
    operation_mode = v.get("operation_mode", "")
    if load_type == LoadType.VFD and not vfd_model:
        # 运行方式标明是变频但无变频器型号，尝试从变频器字段获取
        vfd_raw = v.get("vfd_model", "")
        if vfd_raw and vfd_raw != "/" and vfd_raw != "":
            vfd_model = vfd_raw

    record = CircuitRecord(
        row_number=raw.row_number,
        circuit_id=cid,
        circuit_name=circuit_name,
        load_type=load_type,
        rated_power_kw=rated_power,
        rated_current_a=rated_current,
        breaker_model=breaker_model,
        breaker_poles=breaker_poles,
        breaker_trip_current_a=breaker_trip,
        ct_ratio=ct_ratio if ct_ratio else "",
        cable_type=cable_type,
        cable_section=cable_section,
        vfd_model=vfd_model,
        vfd_power_kw=vfd_power,
        remark=v.get("remark"),
        cabinet_code=v.get("cabinet_code"),
        cabinet_size=v.get("cabinet_size"),
        unit_space=v.get("unit_space"),
        distribution_type=v.get("distribution_type"),
        operation_mode=operation_mode if operation_mode else None,
        breaker_frame_current_a=breaker_frame,
        contactor=v.get("contactor") if v.get("contactor") != "/" else None,
        thermal_relay=v.get("thermal_relay") if v.get("thermal_relay") != "/" else None,
        power_monitoring=v.get("power_monitoring"),
        cable_number=v.get("cable_number"),
    )
    return record, None


def _infer_load_type_from_transposed(v: Dict[str, str]) -> Optional[LoadType]:
    """从转置格式数据推断负荷类型

    优先级：
    1. 显式 load_type 字段
    2. 运行方式映射
    3. 配电形式 + 上下文推断
    """
    # 1. 显式字段
    if "load_type" in v and v["load_type"]:
        lt = LoadType.from_str(v["load_type"])
        if lt:
            return lt

    # 2. 运行方式
    op_mode = v.get("operation_mode", "").strip()
    if op_mode:
        op_lower = op_mode.lower()
        for key, load_type in OPERATION_MODE_TO_LOAD_TYPE.items():
            if key in op_lower or op_lower in key:
                return load_type

    # 3. 配电形式推断
    dist_type = v.get("distribution_type", "").strip().upper()
    if dist_type == "MCC":
        return LoadType.POWER
    elif dist_type == "PC":
        return LoadType.POWER

    # 4. 如果有变频器字段且有值，推断为变频
    vfd_val = v.get("vfd_model", "")
    if vfd_val and vfd_val != "/" and vfd_val != "":
        return LoadType.VFD

    # 无法推断时返回 POWER 作为默认
    logger.warning("无法推断负荷类型，使用默认: 动力")
    return LoadType.POWER


def _infer_poles_from_transposed(v: Dict[str, str], rated_current: float) -> int:
    """从转置格式数据推断断路器极数

    策略:
    1. 显式 breaker_poles 字段
    2. 从线缆规格推断（如 4x...+1x... → 3P, 3x... → 1P/2P）
    3. 从电流大小推断（>63A → 3P, else 1P）
    """
    # 1. 显式字段
    if "breaker_poles" in v and v["breaker_poles"]:
        try:
            return _parse_int(v["breaker_poles"])
        except ValueError:
            pass

    # 2. 从线缆规格推断
    cable_full = v.get("cable_full", "")
    cable_section = ""
    # 提取规格部分（空格后面的部分，如 "4x185+1x95"）
    parts = cable_full.split()
    for p in parts:
        if "x" in p.lower() or "×" in p:
            cable_section = p
            break

    if cable_section:
        section_lower = cable_section.lower().replace("×", "x")
        # 4x...+1x... 或 5x... → 三相 (3P)
        # 3x... → 可能是单相 (1P) 或两相 (2P)
        # 2x... → 单相 (1P)
        if section_lower.startswith("4x") or section_lower.startswith("5x"):
            return 3
        elif section_lower.startswith("3x"):
            return 3  # 默认3P
        elif section_lower.startswith("2x"):
            return 1

    # 3. 从电流大小推断
    if rated_current > 63:
        return 3
    elif rated_current > 32:
        return 3
    else:
        return 3  # 默认3P（工业配电常见）


def _split_cable_full(cable_full: str) -> Tuple[str, str]:
    """拆分线缆型号规格

    例如: "YJV-0.6/1kV 4x185+1x95" → ("YJV-0.6/1kV", "4x185+1x95")
          "BV 3x4" → ("BV", "3x4")
    """
    if not cable_full or cable_full == "/":
        return ("", "")

    cable_full = cable_full.strip()

    # 按空格拆分，第一部分为型号，其余为规格
    parts = cable_full.split(None, 1)
    if len(parts) == 2:
        return (parts[0], parts[1])
    elif len(parts) == 1:
        # 尝试用常见型号前缀拆分
        for prefix in ["YJV-", "YJV", "BV-", "BV", "NH-", "ZR-"]:
            if cable_full.startswith(prefix):
                rest = cable_full[len(prefix):]
                if rest and rest[0].isdigit():
                    return (prefix, rest)
                elif rest and rest[0] == "-":
                    # 如 "YJV-0.6/1kV4x185+1x95"
                    idx = 0
                    for i, c in enumerate(rest):
                        if c.isdigit() and "x" in rest[i:]:
                            idx = i
                            break
                    if idx > 0:
                        return (prefix + rest[:idx], rest[idx:])
        return (cable_full, "")

    return (cable_full, "")


# ============================================================
# 共享校验逻辑（标准格式使用）
# ============================================================

def _validate_row(
    raw: RawRow, seen_ids: set, cfg: ExcelConfig = None
) -> Tuple[Optional[CircuitRecord], Optional[ErrorRecord]]:
    """校验单行数据（标准格式）

    Returns:
        (CircuitRecord or None, ErrorRecord or None)
    """
    v = raw.values

    def err(error_type: str, msg: str):
        return ErrorRecord(
            row_number=raw.row_number,
            circuit_id=v.get("circuit_id", ""),
            error_type=error_type,
            error_message=msg,
            raw_values=dict(v),
        )

    # 必填字段非空
    for field in REQUIRED_FIELDS:
        if field not in v or v[field] == "" or v[field] == "None":
            return None, err("缺必填字段", f"行 {raw.row_number}: 字段 \"{field}\" 为空")

    # 回路编号唯一性
    cid = v["circuit_id"]
    if cid in seen_ids:
        return None, err("编号重复", f"行 {raw.row_number}: 回路编号 \"{cid}\" 重复")

    # 负荷类型枚举
    load_type = LoadType.from_str(v["load_type"])
    if load_type is None:
        valid = [lt.value for lt in LoadType]
        return None, err(
            "非法负荷类型",
            f"行 {raw.row_number}: \"{v['load_type']}\" 不在 {valid} 中",
        )

    # 数值转换
    try:
        rated_power = _parse_float(v["rated_power_kw"])
        rated_current = _parse_float(v["rated_current_a"])
        breaker_poles = _parse_int(v["breaker_poles"])
        breaker_trip = _parse_float(v["breaker_trip_current_a"])
    except ValueError as e:
        return None, err("数值转换失败", f"行 {raw.row_number}: {e}")

    # 数值范围
    if rated_power <= 0:
        return None, err("数值越界", f"行 {raw.row_number}: 额定功率须>0，实际={rated_power}")
    if rated_current <= 0:
        return None, err("数值越界", f"行 {raw.row_number}: 额定电流须>0，实际={rated_current}")
    if breaker_poles not in (1, 2, 3, 4):
        return None, err("数值越界", f"行 {raw.row_number}: 极数须为1-4，实际={breaker_poles}")

    # 变频器一致性
    vfd_model = v.get("vfd_model") or None
    vfd_power = None
    if load_type == LoadType.VFD:
        if not vfd_model:
            return None, err(
                "变频器一致性",
                f"行 {raw.row_number}: 负荷类型为变频但变频器型号为空",
            )
        vfd_power_str = v.get("vfd_power_kw") or None
        if vfd_power_str:
            try:
                vfd_power = _parse_float(vfd_power_str)
            except ValueError:
                vfd_power = None

    record = CircuitRecord(
        row_number=raw.row_number,
        circuit_id=cid,
        circuit_name=v["circuit_name"],
        load_type=load_type,
        rated_power_kw=rated_power,
        rated_current_a=rated_current,
        breaker_model=v["breaker_model"],
        breaker_poles=breaker_poles,
        breaker_trip_current_a=breaker_trip,
        ct_ratio=v["ct_ratio"],
        cable_type=v["cable_type"],
        cable_section=v["cable_section"],
        vfd_model=vfd_model,
        vfd_power_kw=vfd_power,
        remark=v.get("remark"),
    )
    return record, None


# ============================================================
# 数值解析辅助
# ============================================================

def _parse_float(value: str) -> float:
    """解析浮点数，支持去除单位后缀（如 100A -> 100.0）"""
    cleaned = re.sub(r"[^\d.\-eE+]", "", str(value))
    if not cleaned:
        raise ValueError(f"无法转换为数值: \"{value}\"")
    return float(cleaned)


def _parse_int(value: str) -> int:
    """解析整数"""
    cleaned = re.sub(r"[^\d]", "", str(value))
    if not cleaned:
        raise ValueError(f"无法转换为整数: \"{value}\"")
    return int(cleaned)
